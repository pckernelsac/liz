from __future__ import annotations

import os
import re
import secrets

import psycopg2
from psycopg2 import Error as PsycopgError
from psycopg2 import IntegrityError as PsycopgIntegrityError
from psycopg2.extras import RealDictCursor
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Mapping, Optional
from dataclasses import dataclass
import json
import random
from math import ceil
import pytz

import asyncio
import logging

from fastapi import FastAPI, Request, UploadFile
from fastapi.responses import JSONResponse as _BaseJSONResponse, RedirectResponse, FileResponse


class JSONResponse(_BaseJSONResponse):
    """JSONResponse that serializes datetime/date/Decimal objects via str()."""

    def render(self, content) -> bytes:
        return json.dumps(
            content,
            ensure_ascii=False,
            allow_nan=False,
            indent=None,
            separators=(",", ":"),
            default=str,
        ).encode("utf-8")
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from starlette.templating import Jinja2Templates
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from PIL import Image
import openpyxl
from dotenv import load_dotenv
from functools import wraps

# ==== NUEVO: PDF con ReportLab ====
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.colors import HexColor

# Cargar variables de entorno
load_dotenv()

# ✅ CONFIGURACIÓN DE ZONA HORARIA PERÚ (UTC-5)
PERU_TZ = pytz.timezone('America/Lima')

def get_peru_time() -> datetime:
    """Obtiene la fecha y hora actual en la zona horaria de Perú (UTC-5)."""
    return datetime.now(PERU_TZ)

def format_datetime_peru(dt_string: str) -> str:
    """Convierte una fecha UTC a hora de Perú y la formatea."""
    try:
        if not dt_string:
            return ''
        # Si la fecha ya tiene timezone info, convertirla
        if 'T' in dt_string or ' ' in dt_string:
            # Intentar parsear como ISO format
            dt = datetime.fromisoformat(dt_string.replace('Z', '+00:00'))
            if dt.tzinfo is None:
                # Asumir UTC si no tiene timezone
                dt = pytz.utc.localize(dt)
            # Convertir a hora de Perú
            dt_peru = dt.astimezone(PERU_TZ)
            return dt_peru.strftime('%Y-%m-%d %H:%M:%S')
        return dt_string
    except Exception as e:
        return dt_string

# Configuración de la aplicación
@dataclass
class AppConfig:
    """Configuración de la aplicación usando dataclass."""
    SECRET_KEY: str = os.getenv('SECRET_KEY', secrets.token_hex(32))
    UPLOAD_FOLDER: str = 'uploads'
    MAX_CONTENT_LENGTH: int = 16 * 1024 * 1024  # 16MB
    ALLOWED_EXTENSIONS: set[str] = None
    DEBUG: bool = os.getenv('DEBUG', 'True').lower() == 'true'
    HOST: str = os.getenv('HOST', '0.0.0.0')
    PORT: int = int(os.getenv('PORT', '5000'))
    # PostgreSQL: postgresql://user:password@host:5432/dbname
    DATABASE_URL: str = os.getenv(
        'DATABASE_URL',
        'postgresql://postgres:postgres@localhost:5432/sorteo',
    )

    def __post_init__(self):
        if self.ALLOWED_EXTENSIONS is None:
            self.ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'webp'}

@dataclass
class Participante:
    tipo_documento: str
    numero_documento: str
    nombres: str
    apellidos: str
    whatsapp: str
    departamento: str
    comprobante_path: Optional[str] = None
    id: Optional[int] = None
    fecha_registro: Optional[str] = None
    activo: bool = True

@dataclass
class Premio:
    nombre: str
    descripcion: Optional[str] = None
    imagen_path: Optional[str] = None
    orden: int = 0
    id: Optional[int] = None
    activo: bool = True

@dataclass 
class Sorteo:
    nombre: str
    fecha_sorteo: str
    descripcion: Optional[str] = None
    estado: str = 'programado'
    id: Optional[int] = None
    fecha_creacion: Optional[str] = None

# Inicializar configuración
config = AppConfig()

logger = logging.getLogger(__name__)

templates = Jinja2Templates(directory="templates")


def _as_date(value):
    """Formatea un valor de fecha (datetime, date o str 'YYYY-MM-DD...') como 'YYYY-MM-DD'."""
    if value is None or value == '':
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    s = str(value)
    return s[:10]


def _as_time(value):
    """Formatea un valor de fecha como 'HH:MM'. Acepta datetime o str ISO."""
    if value is None or value == '':
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%H:%M')
    s = str(value)
    return s[11:16] if len(s) >= 16 else ''


templates.env.filters['as_date'] = _as_date
templates.env.filters['as_time'] = _as_time


def _missing_url_param(v) -> bool:
    """True si falta id para armar URL (None, '', Jinja Undefined)."""
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    try:
        from jinja2.runtime import Undefined

        if isinstance(v, Undefined):
            return True
    except ImportError:
        pass
    return False


def make_url_for(request: Request):
    """Equivalente mínimo a flask.url_for para las plantillas Jinja existentes."""

    def url_for(endpoint: str, **values):
        base = str(request.base_url).rstrip("/")
        if endpoint == "static":
            filename = values.get("filename", "").lstrip("/")
            return f"{base}/static/{filename}"

        sid = values.get("sorteo_id")
        if _missing_url_param(sid):
            ver_ganadores_sorteo_publico = f"{base}/sorteos"
        else:
            ver_ganadores_sorteo_publico = f"{base}/sorteos/{sid}/ganadores/ver"

        routes = {
            "index": f"{base}/",
            "ver_premios": f"{base}/premios",
            "mis_tickets": f"{base}/mis-tickets",
            "lista_sorteos_publico": f"{base}/sorteos",
            "ver_ganadores_sorteo_publico": ver_ganadores_sorteo_publico,
            "admin": f"{base}/admin",
            "gestion_sorteos": f"{base}/admin/gestion-sorteos",
            "descargar_tickets_pdf": f"{base}/descargar-tickets-pdf",
            "validacion_documentos": f"{base}/admin/validacion-documentos",
            "exportar_participantes": f"{base}/exportar-participantes",
            "participantes_vivo": f"{base}/participantes-vivo",
        }
        return routes.get(endpoint, base + "/")

    return url_for


def template_ctx(request: Request, **extra):
    session = request.session if "session" in request.scope else {}
    return {"request": request, "url_for": make_url_for(request), "session": session, **extra}


def qp_int(request: Request, key: str, default: int) -> int:
    raw = request.query_params.get(key)
    if raw is None:
        return default
    try:
        return int(raw)
    except ValueError:
        return default


def qp_str(request: Request, key: str, default: str = "") -> str:
    v = request.query_params.get(key)
    return default if v is None else str(v)


app = FastAPI(title="Sistema de Sorteos", debug=config.DEBUG)
app.add_middleware(SessionMiddleware, secret_key=config.SECRET_KEY, max_age=3600, same_site="lax")


@app.on_event("startup")
def _startup_init_db() -> None:
    # Inicializa esquema (idempotente: usa CREATE TABLE IF NOT EXISTS).
    # Necesario cuando se arranca con `uvicorn app:app` sin pasar por main().
    try:
        init_database()
    except Exception as exc:
        logger.error(f"Fallo inicializando la BD en startup: {exc}")
        raise

# Configuración de administrador por defecto
ADMIN_CREDENTIALS = {
    'admin': generate_password_hash('admin123'),
    'root': generate_password_hash('root2024'),
    'lorenzo': generate_password_hash('premios123')
}

# Crear directorios necesarios
Path(config.UPLOAD_FOLDER).mkdir(parents=True, exist_ok=True)
Path('static/img').mkdir(parents=True, exist_ok=True)
Path('static').mkdir(parents=True, exist_ok=True)


def _sql_qmarks_to_pct(s: str) -> str:
    """Adapta placeholders estilo SQLite (?) a psycopg2 (%s)."""
    return s.replace('?', '%s')


class PgCursor:
    """Cursor compatible con el patrón sqlite3: execute → fetchone/fetchall; lastrowid tras RETURNING."""

    def __init__(self, raw):
        self._raw = raw
        self.rowcount = 0
        self.lastrowid: Optional[int] = None

    def execute(self, query: str, params: tuple | list | None = None):
        q = _sql_qmarks_to_pct(query)
        self._raw.execute(q, params or ())
        self.rowcount = self._raw.rowcount
        self.lastrowid = None
        if re.search(r'\bRETURNING\b', q, re.IGNORECASE):
            row = self._raw.fetchone()
            if row is not None:
                self.lastrowid = row.get('id') if isinstance(row, dict) else row['id']
        return self

    def fetchone(self):
        return self._raw.fetchone()

    def fetchall(self):
        return self._raw.fetchall()


class PgConnection:
    """Conexión envuelta con execute() como sqlite3."""

    def __init__(self, raw_conn):
        self._raw = raw_conn

    def cursor(self) -> PgCursor:
        return PgCursor(self._raw.cursor(cursor_factory=RealDictCursor))

    def execute(self, query: str, params: tuple | list | None = None) -> PgCursor:
        c = self.cursor()
        return c.execute(query, params)

    def commit(self) -> None:
        self._raw.commit()

    def rollback(self) -> None:
        self._raw.rollback()

    def close(self) -> None:
        self._raw.close()


class DatabaseManager:
    """Gestor de base de datos PostgreSQL con context manager."""

    def __init__(self, dsn: str = config.DATABASE_URL):
        self.dsn = dsn

    def __enter__(self) -> PgConnection:
        self._conn = psycopg2.connect(self.dsn)
        return PgConnection(self._conn)

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self._conn.commit()
        else:
            self._conn.rollback()
        self._conn.close()

class FileManager:
    """Gestor de archivos con validación mejorada."""
    @staticmethod
    def allowed_file(filename: str) -> bool:
        return ('.' in filename and filename.rsplit('.', 1)[1].lower() in config.ALLOWED_EXTENSIONS)

    @staticmethod
    def save_file(file, upload_folder: str = config.UPLOAD_FOLDER) -> str | None:
        if not file or file.filename == '' or not FileManager.allowed_file(file.filename):
            return None

        filename = secure_filename(file.filename)
        timestamp = get_peru_time().strftime('%Y%m%d_%H%M%S_')
        filename = f"{timestamp}{filename}"
        file_path = Path(upload_folder) / filename

        # Guardar archivo
        file.save(file_path)

        # Optimizar imagen si es una imagen
        if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.webp')):
            FileManager.optimize_image(file_path)

        return filename

    @staticmethod
    async def save_upload_file(file: UploadFile | None, upload_folder: str = config.UPLOAD_FOLDER) -> str | None:
        if file is None or not getattr(file, 'filename', None) or not FileManager.allowed_file(file.filename):
            return None
        raw = secure_filename(file.filename)
        timestamp = get_peru_time().strftime('%Y%m%d_%H%M%S_')
        filename = f"{timestamp}{raw}"
        file_path = Path(upload_folder) / filename
        content = await file.read()
        file_path.write_bytes(content)
        if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.webp')):
            FileManager.optimize_image(file_path)
        return filename

    @staticmethod
    def optimize_image(file_path: Path, max_size: tuple[int, int] = (1200, 1200)) -> None:
        try:
            with Image.open(file_path) as img:
                if img.mode in ('RGBA', 'P'):
                    img = img.convert('RGB')
                img.thumbnail(max_size, Image.Resampling.LANCZOS)
                img.save(file_path, optimize=True, quality=85)
        except Exception as e:
            logger.warning(f"No se pudo optimizar la imagen {file_path}: {e}")

def _generar_numero_participacion_unico(conn: PgConnection, max_attempts: int = 1000) -> str:
    """Genera un número único de participación de 6 dígitos con protección contra race conditions."""
    import random
    import time

    # Usar un generador aleatorio más robusto
    secure_random = random.SystemRandom()

    # Intentar generar un número único
    for attempt in range(max_attempts):
        # Para los primeros intentos, usar números aleatorios
        if attempt < 100:
            numero = secure_random.randint(100000, 999999)
        else:
            # Si hay muchos intentos fallidos, combinar timestamp con aleatorio
            base_time = int(time.time() * 1000) % 1000000
            offset = secure_random.randint(0, 999)
            numero = ((base_time + offset) % 900000) + 100000

        numero_str = str(numero)

        # Verificar que no exista
        resultado = conn.execute('SELECT COUNT(*) as count FROM participantes WHERE numero_participacion = ?', (numero_str,)).fetchone()
        if resultado['count'] == 0:
            return numero_str

    # Si no se pudo generar después de muchos intentos, usar secuencia basada en timestamp
    # Esto garantiza unicidad incluso bajo alta carga
    timestamp_micro = str(int(time.time() * 1000000))[-6:]
    return timestamp_micro

def _pg_column_exists(cursor: PgCursor, table: str, column: str) -> bool:
    cursor.execute(
        """
        SELECT 1 FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = ? AND column_name = ?
        """,
        (table, column),
    )
    return cursor.fetchone() is not None


def init_database() -> None:
    with DatabaseManager() as conn:
        cursor = conn.cursor()

        # Orden de creación respetando FKs: premios y sorteos antes que participantes y ganadores.
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS premios (
                id SERIAL PRIMARY KEY,
                nombre TEXT NOT NULL,
                descripcion TEXT,
                imagen_path TEXT,
                orden INTEGER DEFAULT 0,
                cantidad_ganadores INTEGER DEFAULT 1 CHECK(cantidad_ganadores >= 1),
                activo BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS sorteos (
                id SERIAL PRIMARY KEY,
                nombre TEXT NOT NULL,
                fecha_sorteo TIMESTAMP NOT NULL,
                descripcion TEXT,
                estado TEXT DEFAULT 'activo' CHECK(estado IN ('activo', 'cerrado', 'completado', 'cancelado')),
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                fecha_cierre TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS participantes (
                id SERIAL PRIMARY KEY,
                numero_participacion TEXT UNIQUE NOT NULL,
                sorteo_id INTEGER REFERENCES sorteos (id),
                tipo_documento TEXT NOT NULL CHECK(tipo_documento IN ('DNI', 'CE')),
                numero_documento TEXT NOT NULL,
                nombres TEXT NOT NULL,
                apellidos TEXT NOT NULL,
                whatsapp TEXT NOT NULL,
                departamento TEXT NOT NULL,
                comprobante_path TEXT,
                comprobante_estado TEXT DEFAULT 'pendiente' CHECK(comprobante_estado IN ('pendiente', 'aprobado', 'rechazado')),
                comprobante_observaciones TEXT,
                validado_por TEXT,
                fecha_validacion TIMESTAMP,
                fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                activo BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS ganadores (
                id SERIAL PRIMARY KEY,
                sorteo_id INTEGER NOT NULL REFERENCES sorteos (id) ON DELETE CASCADE,
                participante_id INTEGER NOT NULL REFERENCES participantes (id) ON DELETE CASCADE,
                premio_id INTEGER NOT NULL REFERENCES premios (id) ON DELETE CASCADE,
                fecha_ganador TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                notificado BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(sorteo_id, participante_id)
            )
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS departamentos (
                id SERIAL PRIMARY KEY,
                nombre TEXT UNIQUE NOT NULL,
                codigo TEXT UNIQUE,
                activo BOOLEAN DEFAULT TRUE
            )
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS historial_validaciones (
                id SERIAL PRIMARY KEY,
                participante_id INTEGER NOT NULL REFERENCES participantes (id) ON DELETE CASCADE,
                estado TEXT NOT NULL CHECK(estado IN ('aprobado', 'rechazado')),
                observaciones TEXT,
                validado_por TEXT,
                fecha_validacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS historial_ediciones (
                id SERIAL PRIMARY KEY,
                participante_id INTEGER NOT NULL REFERENCES participantes (id) ON DELETE CASCADE,
                campo_editado TEXT NOT NULL,
                valor_anterior TEXT,
                valor_nuevo TEXT,
                editado_por TEXT NOT NULL,
                fecha_edicion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                observaciones TEXT
            )
            """
        )

        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_historial_ediciones_participante ON historial_ediciones(participante_id)"
        )
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_historial_ediciones_fecha ON historial_ediciones(fecha_edicion)")

        if _pg_column_exists(cursor, "participantes", "numero_participacion") is False:
            cursor.execute("ALTER TABLE participantes ADD COLUMN numero_participacion TEXT")
            participantes_existentes = conn.execute(
                "SELECT id FROM participantes WHERE numero_participacion IS NULL"
            ).fetchall()
            for p in participantes_existentes:
                numero_participacion = _generar_numero_participacion_unico(conn)
                cursor.execute(
                    "UPDATE participantes SET numero_participacion = ? WHERE id = ?",
                    (numero_participacion, p["id"]),
                )
            cursor.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_numero_participacion ON participantes(numero_participacion)"
            )

        if _pg_column_exists(cursor, "participantes", "sorteo_id") is False:
            cursor.execute("ALTER TABLE participantes ADD COLUMN sorteo_id INTEGER REFERENCES sorteos (id)")

        cursor.execute("CREATE INDEX IF NOT EXISTS idx_participantes_sorteo ON participantes(sorteo_id)")
        
        # 🔧 CORREGIDO: Ya no se crean sorteos automáticamente al iniciar
        # Los sorteos deben ser creados manualmente por el administrador desde la interfaz web
        # Solo asignar participantes huérfanos si hay un sorteo activo disponible
        sorteo_activo = conn.execute("SELECT * FROM sorteos WHERE estado = 'activo' LIMIT 1").fetchone()
        if sorteo_activo:
            sorteo_activo_id = sorteo_activo['id']
            print(f"[INFO] Sorteo activo existente: ID {sorteo_activo_id} - {sorteo_activo['nombre']}")
            
            # Asignar participantes sin sorteo al sorteo activo
            cursor.execute('UPDATE participantes SET sorteo_id = ? WHERE sorteo_id IS NULL', (sorteo_activo_id,))
            participantes_asignados = cursor.rowcount
            if participantes_asignados > 0:
                print(f"[OK] {participantes_asignados} participantes asignados al sorteo activo")
        else:
            print("[AVISO] No hay sorteo activo. Los administradores deben crear uno desde la interfaz web.")

        departamentos_peru = [
            ('Amazonas', None), ('Áncash', None), ('Apurímac', None), 
            ('Arequipa', None), ('Ayacucho', None), ('Cajamarca', None),
            ('Callao', None), ('Cusco', None), ('Huancavelica', None), 
            ('Huánuco', None), ('Ica', None), ('Junín', None), 
            ('La Libertad', None), ('Lambayeque', None), ('Lima', None), 
            ('Loreto', None), ('Madre de Dios', None), ('Moquegua', None),
            ('Pasco', None), ('Piura', None), ('Puno', None), 
            ('San Martín', None), ('Tacna', None), ('Tumbes', None),
            ('Ucayali', None)
        ]
        for nombre, codigo in departamentos_peru:
            cursor.execute(
                """
                INSERT INTO departamentos (nombre, codigo) VALUES (?, ?)
                ON CONFLICT (nombre) DO NOTHING
                """,
                (nombre, codigo),
            )

        # 🔧 FIX: Verificar si ya existen premios antes de insertar los de ejemplo
        premios_existentes = conn.execute(
            "SELECT COUNT(*) as count FROM premios WHERE activo IS TRUE"
        ).fetchone()
        
        # Solo insertar premios de ejemplo si la tabla está vacía
        if premios_existentes['count'] == 0:
            print("[INFO] Insertando premios de ejemplo...")
            premios_ejemplo = [
                Premio('Gran Premio Principal', 'Premio principal del sorteo - Valor S/1.000.000', None, 1),
                Premio('Segundo Premio', 'Segundo lugar del sorteo - Valor S/500.000', None, 2),
                Premio('Tercer Premio', 'Tercer lugar del sorteo - Valor S/250.000', None, 3),
                Premio('Premio Consolación 1', 'Cuarto lugar - Valor S/100.000', None, 4),
                Premio('Premio Consolación 2', 'Quinto lugar - Valor S/50.000', None, 5)
            ]
            
            for premio in premios_ejemplo:
                cursor.execute('''
                    INSERT INTO premios (nombre, descripcion, imagen_path, orden) 
                    VALUES (?, ?, ?, ?)
                ''', (premio.nombre, premio.descripcion, premio.imagen_path, premio.orden))
            
            print(f"[OK] Se insertaron {len(premios_ejemplo)} premios de ejemplo")
        else:
            print(f"[INFO] Ya existen {premios_existentes['count']} premios, omitiendo insercion de ejemplos")

# Decoradores para manejo de errores
def handle_database_error(func):
    """Decorador para manejar errores de base de datos (sync y async)."""
    if asyncio.iscoroutinefunction(func):

        @wraps(func)
        async def awrapper(*args, **kwargs):
            try:
                return await func(*args, **kwargs)
            except PsycopgIntegrityError as e:
                logger.error(f"Error de integridad en base de datos: {e}")
                return JSONResponse(content={'success': False, 'message': 'Error: Datos duplicados o inválidos'}, status_code=400)
            except PsycopgError as e:
                logger.error(f"Error de base de datos: {e}")
                return JSONResponse(content={'success': False, 'message': 'Error interno del servidor'}, status_code=500)
            except Exception as e:
                logger.error(f"Error inesperado: {e}")
                return JSONResponse(content={'success': False, 'message': 'Error interno del servidor'}, status_code=500)

        awrapper.__name__ = func.__name__
        return awrapper

    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except PsycopgIntegrityError as e:
            logger.error(f"Error de integridad en base de datos: {e}")
            return JSONResponse(content={'success': False, 'message': 'Error: Datos duplicados o inválidos'}, status_code=400)
        except PsycopgError as e:
            logger.error(f"Error de base de datos: {e}")
            return JSONResponse(content={'success': False, 'message': 'Error interno del servidor'}, status_code=500)
        except Exception as e:
            logger.error(f"Error inesperado: {e}")
            return JSONResponse(content={'success': False, 'message': 'Error interno del servidor'}, status_code=500)

    wrapper.__name__ = func.__name__
    return wrapper

# Manejadores de errores globales
@app.exception_handler(413)
async def handle_file_too_large(request: Request, exc: Exception):
    return JSONResponse(content={'success': False, 'message': 'Archivo demasiado grande. Máximo permitido: 16MB'}, status_code=413)


@app.exception_handler(404)
async def handle_not_found(request: Request, exc: Exception):
    try:
        return templates.TemplateResponse('404.html', template_ctx(request), status_code=404)
    except Exception:
        return JSONResponse(content={'success': False, 'message': 'Recurso no encontrado'}, status_code=404)


@app.exception_handler(500)
async def handle_internal_error(request: Request, exc: Exception):
    try:
        return templates.TemplateResponse('500.html', template_ctx(request), status_code=500)
    except Exception:
        return JSONResponse(content={'success': False, 'message': 'Error interno del servidor'}, status_code=500)

# Decoradores de autenticación (compatible con FastAPI: inyecta Request)
def login_required(f):
    if asyncio.iscoroutinefunction(f):
        @wraps(f)
        async def awrapper(request: Request, *args, **kwargs):
            if not request.session.get('admin_logged_in'):
                return RedirectResponse(url='/admin/login', status_code=302)
            return await f(request, *args, **kwargs)
        return awrapper

    @wraps(f)
    def decorated_function(request: Request, *args, **kwargs):
        if not request.session.get('admin_logged_in'):
            return RedirectResponse(url='/admin/login', status_code=302)
        return f(request, *args, **kwargs)
    return decorated_function

def check_admin_credentials(username: str, password: str) -> bool:
    if username in ADMIN_CREDENTIALS:
        return check_password_hash(ADMIN_CREDENTIALS[username], password)
    return False

def get_sorteo_activo() -> Optional[dict]:
    """✅ NUEVO: Obtiene el sorteo activo actual."""
    with DatabaseManager() as conn:
        sorteo = conn.execute("SELECT * FROM sorteos WHERE estado = 'activo' LIMIT 1").fetchone()
        return dict(sorteo) if sorteo else None

# Rutas de la aplicación
@app.get('/')
def index(request: Request) -> str:
    with DatabaseManager() as conn:
        departamentos = conn.execute('SELECT * FROM departamentos WHERE activo IS TRUE ORDER BY nombre').fetchall()
    return templates.TemplateResponse('index.html', template_ctx(request, departamentos=departamentos))

@app.post('/registrar')
@handle_database_error
async def registrar_participante(request: Request):
    """Registrar nuevo participante - Solo 1 ticket, cantidad asignada por admin."""
    form = await request.form()
    # Validar datos del formulario
    required_fields = ['tipo_documento', 'numero_documento', 'nombres', 'apellidos', 'whatsapp', 'departamento']
    form_data = {field: (form.get(field) or '').strip() for field in required_fields}
    
    # Verificar campos obligatorios
    if not all(form_data[field] for field in required_fields):
        return JSONResponse(content={
            'success': False, 
            'message': 'Todos los campos son obligatorios'
        }, status_code=400)
        # Cantidad fija de 1 ticket por registro
    cantidad_tickets = 1
    
    # ✅ NUEVO: Verificar que haya un sorteo activo
    sorteo_activo = get_sorteo_activo()
    if not sorteo_activo:
        return JSONResponse(content={
            'success': False,
            'message': 'No hay ningún sorteo activo en este momento. Por favor, inténtelo más tarde.'
        }, status_code=400)
        # Verificar aceptación de términos
    if not form.get('acepta_terminos'):
        return JSONResponse(content={
            'success': False, 
            'message': 'Debe aceptar los términos y condiciones'
        }, status_code=400)
        # Validar tipo de documento
    if form_data['tipo_documento'] not in ['DNI', 'CE']:
        return JSONResponse(content={
            'success': False, 
            'message': 'Tipo de documento inválido'
        }, status_code=400)
        # Validar número de documento (solo números, mínimo 7 dígitos)
    if not form_data['numero_documento'].isdigit() or len(form_data['numero_documento']) < 7:
        return JSONResponse(content={
            'success': False, 
            'message': 'Número de documento inválido (mínimo 7 dígitos)'
        }, status_code=400)
        # Validar WhatsApp (números, espacios, guiones, paréntesis, +)
    if not re.match(r'^[\d\s\-\(\)\+]+$', form_data['whatsapp']):
        return JSONResponse(content={
            'success': False, 
            'message': 'Número de WhatsApp inválido'
        }, status_code=400)
        # 🔧 VALIDACIÓN DE REGISTROS DUPLICADOS Y PENDIENTES
    with DatabaseManager() as conn:
        # Verificar si el usuario ya tiene tickets PENDIENTES en el sorteo activo
        tickets_pendientes = conn.execute('''
            SELECT COUNT(*) as count FROM participantes 
            WHERE numero_documento = ? 
            AND sorteo_id = ?
            AND comprobante_estado = 'pendiente'
            AND activo IS TRUE
        ''', (form_data['numero_documento'], sorteo_activo['id'])).fetchone()
        
        if tickets_pendientes['count'] > 0:
            return JSONResponse(content={
                'success': False,
                'message': f'Ya tienes {tickets_pendientes["count"]} ticket(s) pendiente(s) de validación. Por favor espera a que el administrador valide tu comprobante antes de realizar un nuevo registro. Puedes consultar el estado de tus tickets con tu número de documento.'
            }, status_code=400)
        # Manejar archivo de comprobante (solo uno para todos los tickets)
    comprobante_path = None
    upload = form.get('comprobante')
    if upload is not None and hasattr(upload, 'filename') and getattr(upload, 'filename', None):
        comprobante_path = await FileManager.save_upload_file(upload)
        if not comprobante_path:
            return JSONResponse(content={
                'success': False,
                'message': 'Error al guardar el comprobante. Por favor, intenta de nuevo.'
            }, status_code=400)
        # Crear objetos participante
    participante_base = {
        'tipo_documento': form_data['tipo_documento'],
        'numero_documento': form_data['numero_documento'],
        'nombres': form_data['nombres'].title(),  # Capitalizar nombres
        'apellidos': form_data['apellidos'].title(),  # Capitalizar apellidos
        'whatsapp': form_data['whatsapp'],
        'departamento': form_data['departamento'],
        'comprobante_path': comprobante_path
    }
    
    ticket_ids = []
    numeros = []  # NUEVO: lista con los números reales de participación
    
    # Guardar en base de datos
    with DatabaseManager() as conn:
        cursor = conn.cursor()

        # Crear todos los tickets solicitados
        for _ in range(cantidad_tickets):
            # Retry logic for race condition protection
            max_retries = 10
            ticket_created = False

            for retry_attempt in range(max_retries):
                try:
                    # Generar número de participación único
                    numero_participacion = _generar_numero_participacion_unico(conn)

                    cursor.execute('''
                        INSERT INTO participantes
                        (numero_participacion, sorteo_id, tipo_documento, numero_documento, nombres, apellidos, whatsapp, departamento, comprobante_path, fecha_registro, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        RETURNING id
                    ''', (
                        numero_participacion,
                        sorteo_activo['id'],
                        participante_base['tipo_documento'],
                        participante_base['numero_documento'],
                        participante_base['nombres'],
                        participante_base['apellidos'],
                        participante_base['whatsapp'],
                        participante_base['departamento'],
                        participante_base['comprobante_path'],
                        get_peru_time().isoformat(),
                        get_peru_time().isoformat(),
                        get_peru_time().isoformat()
                    ))

                    ticket_id = cursor.lastrowid
                    ticket_ids.append(ticket_id)
                    numeros.append(numero_participacion)
                    ticket_created = True
                    break  # Success, exit retry loop

                except PsycopgIntegrityError as e:
                    if 'numero_participacion' in str(e) and retry_attempt < max_retries - 1:
                        # Race condition detected, retry with a new number
                        logger.warning(f"numero_participacion collision (attempt {retry_attempt + 1}/{max_retries}), retrying...")
                        import time
                        time.sleep(0.01 * (retry_attempt + 1))  # Small exponential backoff
                        continue
                    else:
                        # Other integrity error or max retries reached
                        logger.error(f"Error creating ticket after {retry_attempt + 1} attempts: {e}")
                        break

            if not ticket_created:
                logger.error(f"Could not create ticket after {max_retries} attempts")
    
    # Preparar mensaje de respuesta
    message = None
    if len(numeros) >= 1:
        message = '¡Registro exitoso! Tu solicitud ha sido recibida. El administrador verificará tu comprobante y asignará la cantidad de tickets de acuerdo al monto depositado. Te notificaremos cuando tu participación sea confirmada.'
    else:
        message = 'No se pudo completar el registro.'
    
    return JSONResponse(content={
        'success': True if numeros else False, 
        'message': message,
        'ticket_ids': ticket_ids,
        'numeros': numeros,
        'cantidad_registrada': len(ticket_ids)
    }, status_code=201)

@app.get('/premios')
def ver_premios(request: Request) -> str:
    es_admin = bool(request.session.get('admin_logged_in'))
    with DatabaseManager() as conn:
        premios = conn.execute('SELECT * FROM premios WHERE activo IS TRUE ORDER BY orden').fetchall()
    return templates.TemplateResponse('premios.html', template_ctx(request, premios=premios, es_admin=es_admin))

@app.get('/mis-tickets')
def mis_tickets(request: Request) -> str:
    return templates.TemplateResponse('mis_tickets.html', template_ctx(request))

@app.post('/consultar-ticket')
@handle_database_error  
async def consultar_ticket(request: Request):
    """✅ NUEVO: Consulta tickets de sorteos ACTIVOS Y PASADOS con estado de cada sorteo"""
    form = await request.form()
    numero_documento = (form.get('numero_documento') or '').strip()
    if not numero_documento:
        return JSONResponse(content={'success': False, 'message': 'Debe ingresar un número de documento'}, status_code=400)
    with DatabaseManager() as conn:
        # ✅ Traer TODOS los participantes del usuario (activos y pasados)
        participantes_query = conn.execute('''
            SELECT p.*, s.nombre as sorteo_nombre, s.estado as sorteo_estado, s.fecha_sorteo
            FROM participantes p
            LEFT JOIN sorteos s ON p.sorteo_id = s.id
            WHERE p.numero_documento = ? AND p.activo IS TRUE
            ORDER BY p.fecha_registro DESC
        ''', (numero_documento,)).fetchall()

        if not participantes_query:
            return JSONResponse(content={'success': False, 'message': 'No se encontró ningún registro con ese documento'}, status_code=404)
        # Agrupar participantes por sorteo
        sorteos_dict = {}
        for p in participantes_query:
            sorteo_id = p['sorteo_id'] or 0
            if sorteo_id not in sorteos_dict:
                sorteos_dict[sorteo_id] = {
                    'sorteo_id': sorteo_id,
                    'sorteo_nombre': p['sorteo_nombre'] or 'Sin sorteo',
                    'sorteo_estado': p['sorteo_estado'] or 'desconocido',
                    'participantes': []
                }
            sorteos_dict[sorteo_id]['participantes'].append(p)

        # Procesar cada sorteo
        sorteos_info = []
        for sorteo_id, sorteo_data in sorteos_dict.items():
            participantes = sorteo_data['participantes']
            
            # Clasificar por estado de comprobante
            aprobados = [p for p in participantes if p['comprobante_estado'] == 'aprobado']
            pendientes = [p for p in participantes if p['comprobante_path'] and p['comprobante_estado'] == 'pendiente']
            rechazados = [p for p in participantes if p['comprobante_estado'] == 'rechazado']

            # Tickets aprobados
            tickets_aprobados = [p['numero_participacion'] for p in aprobados]

            # Buscar ganadores de este sorteo
            ids_aprobados = [p['id'] for p in aprobados]
            ganadores = []
            if ids_aprobados:
                placeholders = ','.join('?' * len(ids_aprobados))
                ganadores = conn.execute(f'''
                    SELECT g.*, pr.nombre as premio_nombre, pr.orden as premio_orden
                    FROM ganadores g
                    JOIN premios pr ON g.premio_id = pr.id
                    WHERE g.participante_id IN ({placeholders})
                    AND g.sorteo_id = ?
                    ORDER BY pr.orden ASC
                ''', tuple(ids_aprobados) + (sorteo_id,)).fetchall()

            # Estado de participación en este sorteo
            if aprobados:
                estado_participacion = 'activa'
            elif pendientes:
                estado_participacion = 'pendiente'
            elif rechazados:
                estado_participacion = 'rechazada'
            else:
                estado_participacion = 'sin_participacion'

            sorteos_info.append({
                'sorteo_id': sorteo_id,
                'sorteo_nombre': sorteo_data['sorteo_nombre'],
                'sorteo_estado': sorteo_data['sorteo_estado'],
                'es_sorteo_pasado': sorteo_data['sorteo_estado'] != 'activo',
                'tickets': tickets_aprobados,
                'total_tickets': len(tickets_aprobados),
                'estado_participacion': estado_participacion,
                'ganadores': [dict(g) for g in ganadores]
            })

        # Obtener datos básicos del participante (del registro más reciente)
        participante_principal = participantes_query[0]
        participante_data = {
            'nombres': participante_principal['nombres'],
            'apellidos': participante_principal['apellidos'],
            'tipo_documento': participante_principal['tipo_documento'],
            'documento_parcial': f"***{participante_principal['numero_documento'][-3:]}",
            'whatsapp_parcial': f"***{participante_principal['whatsapp'][-4:]}",
            'departamento_parcial': participante_principal['departamento'][:3] + "***",
            'fecha_registro': participante_principal['fecha_registro'],
            'tiene_comprobante': participante_principal['comprobante_path'] is not None,
            'estado_comprobante': participante_principal['comprobante_estado'] if participante_principal['comprobante_path'] else 'sin_comprobante'
        }

        return JSONResponse(content={
            'success': True,
            'participante': participante_data,
            'sorteos': sorteos_info
        }, status_code=200)

@app.get('/admin/login')
def admin_login_get(request: Request):
    return templates.TemplateResponse('login.html', template_ctx(request))


@app.post('/admin/login')
async def admin_login_post(request: Request):
    form = await request.form()
    username = (form.get('username') or '').strip().lower()
    password = form.get('password') or ''
    remember = form.get('remember')

    if not username or not password:
        return JSONResponse(content={'success': False, 'message': 'Usuario y contraseña son obligatorios'}, status_code=400)

    if check_admin_credentials(username, password):
        request.session['admin_logged_in'] = True
        request.session['admin_username'] = username
        request.session['admin_login_time'] = get_peru_time().isoformat()
        if remember:
            request.session['remember'] = True
        return JSONResponse(
            content={'success': True, 'message': f'Bienvenido {username.title()}', 'redirect': '/admin'},
            status_code=200,
        )

    logger.warning(f"Intento de login fallido para usuario: {username} desde IP: {request.client.host if request.client else 'unknown'}")
    return JSONResponse(content={'success': False, 'message': 'Credenciales incorrectas. Verifique usuario y contraseña.'}, status_code=401)


@app.get('/admin/logout')
def admin_logout(request: Request):
    request.session.clear()
    return RedirectResponse(url='/admin/login', status_code=302)

@app.get('/admin')
@login_required
def admin(request: Request) -> str:
    # ✅ Obtener el sorteo activo
    sorteo_activo = get_sorteo_activo()
    sorteo_activo_id = sorteo_activo['id'] if sorteo_activo else None
    
    with DatabaseManager() as conn:
        # 📊 Estadísticas solo del sorteo ACTIVO
        if sorteo_activo_id:
            stats = {
                'total_participantes': conn.execute(
                    "SELECT COUNT(*) as count FROM participantes WHERE activo IS TRUE AND comprobante_estado = 'aprobado' AND sorteo_id = ?",
                    (sorteo_activo_id,)
                ).fetchone()['count'],
                'total_premios': conn.execute('SELECT COUNT(*) as count FROM premios WHERE activo IS TRUE').fetchone()['count'],
                'total_sorteos': conn.execute('SELECT COUNT(*) as count FROM sorteos').fetchone()['count'],
                'total_ganadores': conn.execute('SELECT COUNT(*) as count FROM ganadores').fetchone()['count'],
                'participantes_hoy': conn.execute("""
                    SELECT COUNT(*) as count FROM participantes 
                    WHERE (fecha_registro::date = CURRENT_DATE)
                    AND activo IS TRUE 
                    AND comprobante_estado = 'aprobado'
                    AND sorteo_id = ?
                """, (sorteo_activo_id,)).fetchone()['count']
            }
            participantes_por_depto = conn.execute('''
                SELECT departamento, COUNT(*) as total
                FROM participantes 
                WHERE activo IS TRUE 
                AND comprobante_estado = 'aprobado'
                AND sorteo_id = ?
                GROUP BY departamento 
                ORDER BY total DESC 
                LIMIT 10
            ''', (sorteo_activo_id,)).fetchall()
        else:
            # Si no hay sorteo activo, mostrar ceros
            stats = {
                'total_participantes': 0,
                'total_premios': conn.execute('SELECT COUNT(*) as count FROM premios WHERE activo IS TRUE').fetchone()['count'],
                'total_sorteos': conn.execute('SELECT COUNT(*) as count FROM sorteos').fetchone()['count'],
                'total_ganadores': conn.execute('SELECT COUNT(*) as count FROM ganadores').fetchone()['count'],
                'participantes_hoy': 0
            }
            participantes_por_depto = []
    
    # Datos de paginación para la carga inicial
    per_page = 10
    total_participantes = stats['total_participantes']
    total_pages = ceil(total_participantes / per_page)
    
    with DatabaseManager() as conn:
        participantes_recientes = conn.execute('''
            SELECT * FROM participantes 
            WHERE activo IS TRUE 
            ORDER BY fecha_registro DESC 
            LIMIT ? OFFSET ?
        ''', (per_page, 0)).fetchall()

    return templates.TemplateResponse(
        'admin.html',
        template_ctx(
            request,
            stats=stats,
            participantes_por_depto=participantes_por_depto,
            participantes_recientes=participantes_recientes,
            total_pages=total_pages,
            current_page=1,
            per_page=per_page,
        ),
    )
                           
@app.get('/admin/participantes-recientes')
@login_required
def get_participantes_recientes(request: Request):
    page = qp_int(request, 'page', 1)
    per_page = qp_int(request, 'per_page', 10)
    offset = (page - 1) * per_page
    
    # ✅ Obtener sorteo activo
    sorteo_activo = get_sorteo_activo()
    sorteo_activo_id = sorteo_activo['id'] if sorteo_activo else None
    
    with DatabaseManager() as conn:
        # Mostrar todos los participantes (de todos los sorteos) en la tabla
        participantes = conn.execute('''
            SELECT * FROM participantes 
            WHERE activo IS TRUE 
            ORDER BY fecha_registro DESC 
            LIMIT ? OFFSET ?
        ''', (per_page, offset)).fetchall()
        
        # Pero el total solo del sorteo ACTIVO para paginación
        if sorteo_activo_id:
            total_participantes = conn.execute(
                "SELECT COUNT(*) as count FROM participantes WHERE activo IS TRUE AND comprobante_estado = 'aprobado' AND sorteo_id = ?",
                (sorteo_activo_id,)
            ).fetchone()['count']
        else:
            total_participantes = 0
        
    return JSONResponse(content={
        'success': True,
        'participantes': [dict(p) for p in participantes],
        'total_participantes': total_participantes,
        'per_page': per_page,
        'total_pages': ceil(total_participantes / per_page) if total_participantes > 0 else 1
    }, status_code=200)

@app.post('/realizar-sorteo')
@handle_database_error
def realizar_sorteo(request: Request):
    """Realizar sorteo solo con participantes que tienen comprobante aprobado del sorteo ACTIVO."""
    with DatabaseManager() as conn:
        # ✅ CORREGIDO: Solo participantes con comprobante aprobado del sorteo ACTIVO
        participantes = conn.execute('''
            SELECT p.* FROM participantes p
            INNER JOIN sorteos s ON p.sorteo_id = s.id
            WHERE p.activo IS TRUE 
            AND p.comprobante_estado = 'aprobado'
            AND p.comprobante_path IS NOT NULL
            AND s.estado = 'activo'
        ''').fetchall()
        
        if len(participantes) == 0:
            return JSONResponse(content={
                'success': False, 
                'message': 'No hay participantes con comprobantes aprobados para realizar el sorteo. Por favor, valide los comprobantes primero.'
            }, status_code=400)

        premios = conn.execute('SELECT * FROM premios WHERE activo IS TRUE ORDER BY orden').fetchall()
        if len(premios) == 0:
            return JSONResponse(content={'success': False, 'message': 'No hay premios configurados'}, status_code=400)

        cursor = conn.cursor()
        peru_now = get_peru_time()
        sorteo = Sorteo(
            nombre=f'Sorteo Automático {peru_now.strftime("%Y-%m-%d %H:%M")}',
            fecha_sorteo=peru_now.isoformat(),
            descripcion='Sorteo automático realizado desde el panel administrativo',
            estado='completado'
        )
        cursor.execute(
            '''INSERT INTO sorteos (nombre, fecha_sorteo, descripcion, estado, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?) RETURNING id''',
            (sorteo.nombre, sorteo.fecha_sorteo, sorteo.descripcion, sorteo.estado, peru_now.isoformat(), peru_now.isoformat()),
        )
        sorteo_id = cursor.lastrowid

        secure_random = random.SystemRandom()
        participantes_disponibles = list(participantes)
        ganadores_seleccionados = []

        for premio in premios[:min(len(premios), len(participantes_disponibles))]:
            if participantes_disponibles:
                ganador = secure_random.choice(participantes_disponibles)
                participantes_disponibles.remove(ganador)
                cursor.execute('INSERT INTO ganadores (sorteo_id, participante_id, premio_id) VALUES (?, ?, ?)', 
                               (sorteo_id, ganador['id'], premio['id']))
                ganadores_seleccionados.append({
                    'premio': premio['nombre'],
                    'premio_orden': premio['orden'],
                    'ganador': f"{ganador['nombres']} {ganador['apellidos']}",
                    'numero_ticket': ganador['numero_participacion'],
                    'documento': f"{ganador['tipo_documento']}: {ganador['numero_documento']}",
                    'departamento': ganador['departamento'],
                    'whatsapp': ganador['whatsapp']
                })
    
    return JSONResponse(content={
        'success': True, 
        'message': f'¡Sorteo realizado exitosamente con {len(ganadores_seleccionados)} ganadores!', 
        'sorteo_id': sorteo_id, 
        'ganadores': ganadores_seleccionados,
        'total_participantes': len(participantes),
        'fecha_sorteo': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }, status_code=201)

@app.post('/validar-comprobante')
@handle_database_error
async def validar_comprobante(request: Request):
    try:
        data = await request.json()
    except Exception:
        data = {}
    participante_id = data.get('participante_id')
    estado = data.get('estado')  # 'aprobado' o 'rechazado'
    observaciones = data.get('observaciones', '')
    validado_por = data.get('validado_por', request.session.get('admin_username', 'Administrador'))

    if not participante_id or estado not in ['aprobado', 'rechazado']:
        return JSONResponse(content={'success': False, 'message': 'Datos de validación inválidos'}, status_code=400)
    with DatabaseManager() as conn:
        cursor = conn.cursor()
        participante = conn.execute('SELECT * FROM participantes WHERE id = ?', (participante_id,)).fetchone()
        if not participante:
            return JSONResponse(content={'success': False, 'message': 'Participante no encontrado'}, status_code=404)

        peru_now = get_peru_time().isoformat()
        # La UI agrupa tickets por (numero_documento, comprobante_path), por lo que una validación
        # individual debe afectar todo el grupo (todos los tickets del mismo comprobante).
        grupo = conn.execute('''
            SELECT id FROM participantes
            WHERE numero_documento = ?
              AND comprobante_path IS NOT DISTINCT FROM ?
              AND activo IS TRUE
            ORDER BY id
        ''', (participante['numero_documento'], participante['comprobante_path'])).fetchall()
        afectados = [row['id'] for row in grupo]

        for pid in afectados:
            cursor.execute('''
                UPDATE participantes
                SET comprobante_estado = ?,
                    comprobante_observaciones = ?,
                    validado_por = ?,
                    fecha_validacion = ?,
                    updated_at = ?
                WHERE id = ?
            ''', (estado, observaciones, validado_por, peru_now, peru_now, pid))

            cursor.execute('''
                INSERT INTO historial_validaciones
                (participante_id, estado, observaciones, validado_por, fecha_validacion)
                VALUES (?, ?, ?, ?, ?)
            ''', (pid, estado, observaciones, validado_por, peru_now))

    return JSONResponse(content={
        'success': True,
        'message': f'Comprobante {estado} exitosamente ({len(afectados)} ticket(s))',
        'tickets_afectados': afectados,
    }, status_code=200)

@app.post('/validar-comprobantes-lote')
@handle_database_error
async def validar_comprobantes_lote(request: Request):
    try:
        data = await request.json()
    except Exception:
        data = {}
    participantes_ids = data.get('participantes_ids', [])
    estado = data.get('estado')
    observaciones = data.get('observaciones', '')
    validado_por = data.get('validado_por', request.session.get('admin_username', 'Administrador'))

    if not participantes_ids or estado not in ['aprobado', 'rechazado']:
        return JSONResponse(content={'success': False, 'message': 'Datos de validación inválidos'}, status_code=400)

    procesados, errores = 0, 0
    peru_now = get_peru_time().isoformat()
    with DatabaseManager() as conn:
        cursor = conn.cursor()
        for participante_id in participantes_ids:
            try:
                participante = conn.execute('SELECT * FROM participantes WHERE id = ?', (participante_id,)).fetchone()
                if participante and participante['comprobante_path']:
                    cursor.execute('''
                        UPDATE participantes 
                        SET comprobante_estado = ?, 
                            comprobante_observaciones = ?,
                            validado_por = ?,
                            fecha_validacion = ?,
                            updated_at = ?
                        WHERE id = ?
                    ''', (estado, observaciones, validado_por, peru_now, peru_now, participante_id))
                    cursor.execute('''
                        INSERT INTO historial_validaciones 
                        (participante_id, estado, observaciones, validado_por, fecha_validacion)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (participante_id, estado, observaciones, validado_por, peru_now))
                    procesados += 1
                else:
                    errores += 1
            except Exception as e:
                logger.error(f"Error al procesar participante {participante_id}: {e}")
                errores += 1
    return JSONResponse(content={'success': True if procesados > 0 else False, 'message': f'Se procesaron {procesados} comprobantes correctamente. {errores} errores.', 'procesados': procesados, 'errores': errores}, status_code=200)

@app.get('/estadisticas-validaciones')
@login_required
def estadisticas_validaciones(request: Request):
    with DatabaseManager() as conn:
        stats = {
            'pendientes': conn.execute("SELECT COUNT(*) as count FROM participantes WHERE comprobante_path IS NOT NULL AND comprobante_estado = 'pendiente'").fetchone()['count'],
            'aprobados': conn.execute("SELECT COUNT(*) as count FROM participantes WHERE comprobante_estado = 'aprobado'").fetchone()['count'],
            'rechazados': conn.execute("SELECT COUNT(*) as count FROM participantes WHERE comprobante_estado = 'rechazado'").fetchone()['count'],
            'sin_comprobante': conn.execute("SELECT COUNT(*) as count FROM participantes WHERE comprobante_path IS NULL").fetchone()['count']
        }
        validaciones_por_dia = conn.execute('''
            SELECT 
                (fecha_validacion::date) as fecha,
                COUNT(CASE WHEN comprobante_estado = 'aprobado' THEN 1 END) as aprobados,
                COUNT(CASE WHEN comprobante_estado = 'rechazado' THEN 1 END) as rechazados
            FROM participantes
            WHERE fecha_validacion IS NOT NULL 
                AND (fecha_validacion::date) >= (CURRENT_DATE - INTERVAL '7 days')
            GROUP BY (fecha_validacion::date)
            ORDER BY fecha DESC
        ''').fetchall()
        validaciones_por_admin = conn.execute('''
            SELECT 
                validado_por,
                COUNT(*) as total,
                COUNT(CASE WHEN comprobante_estado = 'aprobado' THEN 1 END) as aprobados,
                COUNT(CASE WHEN comprobante_estado = 'rechazado' THEN 1 END) as rechazados
            FROM participantes
            WHERE validado_por IS NOT NULL
            GROUP BY validado_por
            ORDER BY total DESC
        ''').fetchall()
    return JSONResponse(content={'success': True, 'stats': stats, 'validaciones_por_dia': [dict(v) for v in validaciones_por_dia], 'validaciones_por_admin': [dict(v) for v in validaciones_por_admin]}, status_code=200)

@app.get('/ver-comprobante/{filename:path}')
def ver_comprobante(request: Request, filename: str):
    comprobante_path = Path(config.UPLOAD_FOLDER) / filename
    if not comprobante_path.exists():
        return JSONResponse(content={'success': False, 'message': 'Archivo de comprobante no encontrado'}, status_code=404)

    return FileResponse(comprobante_path)

@app.get('/comprobantes-pendientes')
@login_required
def comprobantes_pendientes(request: Request):
    # Obtener parámetros de paginación
    page = qp_int(request, 'page', 1)
    per_page = qp_int(request, 'per_page', 10)
    search = qp_str(request, 'search', '')
    estado = qp_str(request, 'estado', '')
    departamento = qp_str(request, 'departamento', '')
    
    # Validar parámetros
    if page < 1:
        page = 1
    if per_page < 5 or per_page > 100:
        per_page = 10
    
    offset = (page - 1) * per_page
    
    # Construir la consulta con filtros
    where_conditions = ["p.comprobante_path IS NOT NULL", "p.activo IS TRUE"]
    params = []
    
    # Filtro de búsqueda
    if search:
        where_conditions.append("(p.nombres LIKE ? OR p.apellidos LIKE ? OR p.numero_documento LIKE ? OR p.whatsapp LIKE ?)")
        search_param = f"%{search}%"
        params.extend([search_param, search_param, search_param, search_param])
    
    # Filtro de estado
    if estado and estado != 'todos':
        if estado == 'sin_documento':
            where_conditions[0] = "p.comprobante_path IS NULL"
        else:
            where_conditions.append("p.comprobante_estado = ?")
            params.append(estado)
    
    # Filtro de departamento
    if departamento:
        where_conditions.append("p.departamento = ?")
        params.append(departamento)
    
    where_clause = " AND ".join(where_conditions)
    
    with DatabaseManager() as conn:
        # Consulta principal con paginación
        query = f'''
            SELECT 
                MIN(p.id) as id,
                p.numero_documento,
                MAX(p.tipo_documento) as tipo_documento,
                MAX(p.nombres) as nombres,
                MAX(p.apellidos) as apellidos,
                MAX(p.whatsapp) as whatsapp,
                MAX(p.departamento) as departamento,
                MAX(p.comprobante_path) as comprobante_path,
                MAX(p.comprobante_estado) as comprobante_estado,
                MAX(p.comprobante_observaciones) as comprobante_observaciones,
                MAX(p.fecha_registro) as fecha_registro,
                MAX(p.fecha_validacion) as fecha_validacion,
                MAX(p.validado_por) as validado_por,
                COUNT(p.id) as cantidad_tickets,
                string_agg(p.id::text, ',' ORDER BY p.id) as ids_tickets,
                CASE 
                    WHEN MAX(p.comprobante_estado) = 'pendiente' THEN 'Pendiente'
                    WHEN MAX(p.comprobante_estado) = 'aprobado' THEN 'Aprobado' 
                    WHEN MAX(p.comprobante_estado) = 'rechazado' THEN 'Rechazado'
                END as estado_texto
            FROM participantes p
            WHERE {where_clause}
            GROUP BY p.numero_documento, p.comprobante_path
            ORDER BY 
                CASE WHEN MAX(p.comprobante_estado) = 'pendiente' THEN 1 ELSE 2 END,
                MAX(p.fecha_registro) DESC
            LIMIT ? OFFSET ?
        '''
        
        pendientes = conn.execute(query, params + [per_page, offset]).fetchall()
        
        # Consulta para contar el total (sin GROUP BY para contar correctamente)
        count_query = f'''
            SELECT COUNT(DISTINCT p.numero_documento || '-' || COALESCE(p.comprobante_path, 'sin-comprobante')) as total
            FROM participantes p
            WHERE {where_clause}
        '''
        
        total_count = conn.execute(count_query, params).fetchone()['total']
        
    # Calcular información de paginación
    total_pages = ceil(total_count / per_page)
    has_prev = page > 1
    has_next = page < total_pages
    
    return JSONResponse(content={
        'success': True, 
        'comprobantes': [dict(p) for p in pendientes],
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': total_count,
            'pages': total_pages,
            'has_prev': has_prev,
            'has_next': has_next,
            'prev_num': page - 1 if has_prev else None,
            'next_num': page + 1 if has_next else None
        }
    }, status_code=200)

@app.get('/exportar-participantes')
def exportar_participantes(request: Request):
    try:
        with DatabaseManager() as conn:
            participantes = conn.execute('''
                SELECT p.*, 
                       CASE WHEN p.comprobante_path IS NOT NULL THEN 'Sí' ELSE 'No' END as tiene_comprobante,
                       CASE 
                           WHEN p.comprobante_estado = 'pendiente' THEN 'Pendiente'
                           WHEN p.comprobante_estado = 'aprobado' THEN 'Aprobado' 
                           WHEN p.comprobante_estado = 'rechazado' THEN 'Rechazado'
                           ELSE 'N/A'
                       END as estado_comprobante
                FROM participantes p 
                WHERE p.activo IS TRUE 
                ORDER BY p.fecha_registro DESC
            ''').fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Participantes"
        headers = ['ID', 'Tipo Documento', 'Número Documento', 'Nombres', 'Apellidos', 'WhatsApp', 'Departamento', 'Tiene Comprobante', 'Estado Comprobante', 'Validado Por', 'Fecha Registro']
        ws.append(headers)
        for p in participantes:
            ws.append([
                p['id'], p['tipo_documento'], p['numero_documento'], 
                p['nombres'], p['apellidos'], p['whatsapp'], 
                p['departamento'], p['tiene_comprobante'], p['estado_comprobante'],
                p['validado_por'] or 'N/A',
                (p['fecha_registro'].strftime('%Y-%m-%d %H:%M:%S')
                 if hasattr(p['fecha_registro'], 'strftime')
                 else (str(p['fecha_registro'])[:19] if p['fecha_registro'] else ''))
            ])
        filename = f"participantes_{get_peru_time().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = Path('static') / filename
        wb.save(filepath)
        return FileResponse(filepath, filename=filename, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"Error al exportar: {e}")
        return JSONResponse(content={'success': False, 'message': 'Error al exportar datos'}, status_code=500)

@app.get('/exportar-validaciones')
@login_required
def exportar_validaciones(request: Request):
    try:
        with DatabaseManager() as conn:
            validaciones = conn.execute('''
                SELECT 
                    h.id,
                    p.nombres || ' ' || p.apellidos as participante,
                    p.tipo_documento || ': ' || p.numero_documento as documento,
                    p.whatsapp,
                    p.departamento,
                    h.estado,
                    h.observaciones,
                    h.validado_por,
                    h.fecha_validacion
                FROM historial_validaciones h
                JOIN participantes p ON h.participante_id = p.id
                ORDER BY h.fecha_validacion DESC
            ''').fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial Validaciones"
        headers = ['ID', 'Participante', 'Documento', 'WhatsApp', 'Departamento', 'Estado', 'Observaciones', 'Validado Por', 'Fecha Validación']
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for v in validaciones:
            row = [
                v['id'], v['participante'], v['documento'], v['whatsapp'], v['departamento'],
                v['estado'].title(), v['observaciones'] or 'Sin observaciones', v['validado_por'],
                (v['fecha_validacion'].strftime('%Y-%m-%d %H:%M:%S')
                 if hasattr(v['fecha_validacion'], 'strftime')
                 else (str(v['fecha_validacion'])[:19] if v['fecha_validacion'] else ''))
            ]
            ws.append(row)
            row_num = ws.max_row
            if v['estado'] == 'aprobado':
                ws.cell(row=row_num, column=6).fill = openpyxl.styles.PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            elif v['estado'] == 'rechazado':
                ws.cell(row=row_num, column=6).fill = openpyxl.styles.PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        filename = f"validaciones_{get_peru_time().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = Path('static') / filename
        wb.save(filepath)
        return FileResponse(filepath, filename=filename, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"Error al exportar validaciones: {e}")
        return JSONResponse(content={'success': False, 'message': 'Error al exportar validaciones'}, status_code=500)

@app.get('/historial-validaciones')
@login_required
def historial_validaciones(request: Request):
    page = qp_int(request, 'page', 1)
    per_page = qp_int(request, 'per_page', 50)
    filtro_estado = qp_str(request, 'estado', '')
    filtro_validador = qp_str(request, 'validador', '')
    fecha_desde = qp_str(request, 'fecha_desde', '')
    fecha_hasta = qp_str(request, 'fecha_hasta', '')

    with DatabaseManager() as conn:
        query = '''
            SELECT 
                h.*,
                p.nombres || ' ' || p.apellidos as participante_nombre,
                p.tipo_documento,
                p.numero_documento,
                p.whatsapp,
                p.departamento
            FROM historial_validaciones h
            JOIN participantes p ON h.participante_id = p.id
            WHERE 1=1
        '''
        params = []
        if filtro_estado:
            query += ' AND h.estado = ?'
            params.append(filtro_estado)
        if filtro_validador:
            query += ' AND h.validado_por = ?'
            params.append(filtro_validador)
        if fecha_desde:
            query += ' AND (h.fecha_validacion::date) >= ?::date'
            params.append(fecha_desde)
        if fecha_hasta:
            query += ' AND (h.fecha_validacion::date) <= ?::date'
            params.append(fecha_hasta)
        count_query = f'SELECT COUNT(*) as total FROM ({query})'
        total = conn.execute(count_query, params).fetchone()['total']
        query += ' ORDER BY h.fecha_validacion DESC LIMIT ? OFFSET ?'
        params.extend([per_page, (page - 1) * per_page])
        validaciones = conn.execute(query, params).fetchall()
        validadores = conn.execute('SELECT DISTINCT validado_por FROM historial_validaciones WHERE validado_por IS NOT NULL ORDER BY validado_por').fetchall()
    return JSONResponse(content={'success': True, 'validaciones': [dict(v) for v in validaciones], 'total': total, 'page': page, 'per_page': per_page, 'pages': (total + per_page - 1) // per_page, 'validadores': [v['validado_por'] for v in validadores]}, status_code=200)

@app.get('/admin/validacion-documentos')
@login_required
def validacion_documentos(request: Request):
    with DatabaseManager() as conn:
        departamentos = conn.execute('SELECT * FROM departamentos WHERE activo IS TRUE ORDER BY nombre').fetchall()
    return templates.TemplateResponse('validacion_documentos.html', template_ctx(request, departamentos=departamentos))

@app.get('/admin/gestion-sorteos')
@login_required
def gestion_sorteos(request: Request):
    """✅ NUEVO: Página de gestión de sorteos."""
    return templates.TemplateResponse('gestion_sorteos.html', template_ctx(request))

# ====== RUTAS API PARA GESTIÓN DE PREMIOS ======
@app.post('/admin/premio')
@login_required
@handle_database_error
async def crear_premio(request: Request):
    data = await request.json()
    nombre = data.get('nombre', '').strip()
    descripcion = data.get('descripcion', '').strip()
    orden = data.get('orden', 0)
    cantidad_ganadores = data.get('cantidad_ganadores', 1)  # ✅ NUEVO
    
    if not nombre:
        return JSONResponse(content={'success': False, 'message': 'El nombre del premio es obligatorio'}, status_code=400)
    if not isinstance(orden, int) or orden < 1:
        return JSONResponse(content={'success': False, 'message': 'El orden debe ser un número mayor a 0'}, status_code=400)
    # ✅ NUEVO: Validar cantidad de ganadores
    if not isinstance(cantidad_ganadores, int) or cantidad_ganadores < 1:
        return JSONResponse(content={'success': False, 'message': 'La cantidad de ganadores debe ser al menos 1'}, status_code=400)
    with DatabaseManager() as conn:
        cursor = conn.cursor()
        # ✅ MEJORADO: Incluir cantidad_ganadores
        cursor.execute(
            'INSERT INTO premios (nombre, descripcion, orden, cantidad_ganadores) VALUES (?, ?, ?, ?) RETURNING id',
            (nombre, descripcion, orden, cantidad_ganadores),
        )
        premio_id = cursor.lastrowid

    return JSONResponse(content={'success': True, 'message': 'Premio creado exitosamente', 'premio_id': premio_id}, status_code=201)

@app.get('/admin/premios')
@login_required
@handle_database_error
def listar_premios(request: Request):
    """Lista todos los premios activos del sistema."""
    with DatabaseManager() as conn:
        premios = conn.execute('SELECT * FROM premios WHERE activo IS TRUE ORDER BY orden').fetchall()
    
    return JSONResponse(content={
        'success': True,
        'premios': [dict(p) for p in premios]
    }, status_code=200)

@app.get('/admin/premio/{premio_id}')
@login_required
@handle_database_error
def obtener_premio(request: Request, premio_id: int):
    with DatabaseManager() as conn:
        premio = conn.execute('SELECT * FROM premios WHERE id = ? AND activo IS TRUE', (premio_id,)).fetchone()
        if not premio:
            return JSONResponse(content={'success': False, 'message': 'Premio no encontrado'}, status_code=404)
        return JSONResponse(content={'success': True, 'premio': dict(premio)}, status_code=200)

@app.put('/admin/premio/{premio_id}')
@login_required
@handle_database_error
async def actualizar_premio(request: Request, premio_id: int):
    data = await request.json()
    nombre = data.get('nombre', '').strip()
    descripcion = data.get('descripcion', '').strip()
    orden = data.get('orden', 0)
    cantidad_ganadores = data.get('cantidad_ganadores', 1)  # ✅ NUEVO
    
    if not nombre:
        return JSONResponse(content={'success': False, 'message': 'El nombre del premio es obligatorio'}, status_code=400)
    if not isinstance(orden, int) or orden < 1:
        return JSONResponse(content={'success': False, 'message': 'El orden debe ser un número mayor a 0'}, status_code=400)
    # ✅ NUEVO: Validar cantidad de ganadores
    if not isinstance(cantidad_ganadores, int) or cantidad_ganadores < 1:
        return JSONResponse(content={'success': False, 'message': 'La cantidad de ganadores debe ser al menos 1'}, status_code=400)
    with DatabaseManager() as conn:
        cursor = conn.cursor()
        premio_existente = conn.execute('SELECT * FROM premios WHERE id = ? AND activo IS TRUE', (premio_id,)).fetchone()
        if not premio_existente:
            return JSONResponse(content={'success': False, 'message': 'Premio no encontrado'}, status_code=404)
        # ✅ MEJORADO: Incluir cantidad_ganadores
        cursor.execute('UPDATE premios SET nombre = ?, descripcion = ?, orden = ?, cantidad_ganadores = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
                      (nombre, descripcion, orden, cantidad_ganadores, premio_id))
    
    return JSONResponse(content={'success': True, 'message': 'Premio actualizado exitosamente'}, status_code=200)

@app.delete('/admin/premio/{premio_id}')
@login_required
@handle_database_error
def eliminar_premio(request: Request, premio_id: int):
    with DatabaseManager() as conn:
        cursor = conn.cursor()
        premio_existente = conn.execute('SELECT * FROM premios WHERE id = ? AND activo IS TRUE', (premio_id,)).fetchone()
        if not premio_existente:
            return JSONResponse(content={'success': False, 'message': 'Premio no encontrado'}, status_code=404)
        # Verificar si hay ganadores asociados a este premio
        ganadores = conn.execute('SELECT COUNT(*) as count FROM ganadores WHERE premio_id = ?', (premio_id,)).fetchone()
        if ganadores['count'] > 0:
            return JSONResponse(content={'success': False, 'message': 'No se puede eliminar el premio porque ya tiene ganadores asignados'}, status_code=400)
        cursor.execute('UPDATE premios SET activo = FALSE, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (premio_id,))
    
    return JSONResponse(content={'success': True, 'message': 'Premio eliminado exitosamente'}, status_code=200)

@app.post('/admin/premio/{premio_id}/imagen')
@login_required
@handle_database_error
async def subir_imagen_premio(request: Request, premio_id: int):
    with DatabaseManager() as conn:
        premio_existente = conn.execute('SELECT * FROM premios WHERE id = ? AND activo IS TRUE', (premio_id,)).fetchone()
        if not premio_existente:
            return JSONResponse(content={'success': False, 'message': 'Premio no encontrado'}, status_code=404)

    form = await request.form()
    file = form.get('imagen')
    if file is None or not getattr(file, 'filename', None):
        return JSONResponse(content={'success': False, 'message': 'No se ha seleccionado ninguna imagen'}, status_code=400)
    if not FileManager.allowed_file(file.filename):
        return JSONResponse(content={'success': False, 'message': 'Formato de archivo no permitido. Use PNG, JPG, JPEG, GIF o WEBP'}, status_code=400)

    # Guardar imagen en la carpeta static/img
    imagen_path = await FileManager.save_upload_file(file, 'static/img')
    if not imagen_path:
        return JSONResponse(content={'success': False, 'message': 'Error al guardar la imagen'}, status_code=500)
    with DatabaseManager() as conn:
        cursor = conn.cursor()
        premio_existente = conn.execute('SELECT * FROM premios WHERE id = ? AND activo IS TRUE', (premio_id,)).fetchone()
        # Eliminar imagen anterior si existe
        if premio_existente['imagen_path']:
            old_image_path = Path('static/img') / premio_existente['imagen_path']
            if old_image_path.exists():
                try:
                    old_image_path.unlink()
                except Exception as e:
                    logger.warning(f"No se pudo eliminar imagen anterior: {e}")

        cursor.execute('UPDATE premios SET imagen_path = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
                      (imagen_path, premio_id))

    return JSONResponse(content={'success': True, 'message': 'Imagen actualizada exitosamente'}, status_code=200)

# ====== NUEVO: Generación de Tickets PDF para Admin.html ======
def _build_tickets_table(participantes: list[Mapping[str, Any]]) -> Table:
    from reportlab.lib.styles import ParagraphStyle
    TICKETS_POR_FILA = 6
    data = []
    row = []
    styles = getSampleStyleSheet()
    ticket_style = ParagraphStyle(name='Ticket', parent=styles['Normal'], fontSize=7, alignment=TA_CENTER, leading=9)

    for p in participantes:
        ticket_content = f'<font name="Helvetica-Bold" size=13># {p["numero_participacion"]}</font><br/>{p["nombres"]} {p["apellidos"]}<br/>{p["tipo_documento"]}: {p["numero_documento"]}<br/>Cel: {p["whatsapp"]}'
        row.append(Paragraph(ticket_content, ticket_style))
        if len(row) == TICKETS_POR_FILA:
            data.append(row)
            row = []

    if row:
        while len(row) < TICKETS_POR_FILA:
            row.append(Paragraph(" ", ticket_style))  # Relleno para fila incompleta
        data.append(row)

    col_widths = [28 * mm] * TICKETS_POR_FILA  # 6 cols × 28 mm = 168 mm (entra en A4 útil 170 mm)
    row_heights = [28 * mm] * len(data)  # 9 filas × 28 mm ≈ 252 mm → 54 tickets por hoja A4

    table = Table(data, colWidths=col_widths, rowHeights=row_heights)
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Bordes para recortar
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),  # Fondo blanco
    ]))
    return table

@app.api_route('/generar-tickets-pdf', methods=['GET', 'POST'])
@login_required
def generar_tickets_pdf(request: Request):
    """Genera y guarda un PDF con los tickets APROBADOS (lista de participantes)."""
    try:
        with DatabaseManager() as conn:
            # ✅ CORREGIDO: Solo participantes con comprobante aprobado del sorteo ACTIVO
            participantes = conn.execute('''
                SELECT p.* FROM participantes p
                INNER JOIN sorteos s ON p.sorteo_id = s.id
                WHERE p.activo IS TRUE 
                AND p.comprobante_estado = 'aprobado'
                AND p.comprobante_path IS NOT NULL
                AND s.estado = 'activo'
                ORDER BY p.numero_participacion ASC
            ''').fetchall()

        if not participantes:
            return JSONResponse(content={
                'success': False,
                'message': 'No hay participantes con comprobantes aprobados para generar PDF'
            }, status_code=400)

        doc_path = Path('static') / 'tickets_participantes.pdf'
        doc = SimpleDocTemplate(str(doc_path), pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        elements = []
        title = Paragraph('Listado de Tickets Aprobados - Rapi Rifa', styles['Title'])
        info = Paragraph(f'Generado: {get_peru_time().strftime("%Y-%m-%d %H:%M:%S")} | Total: {len(participantes)} tickets', styles['Normal'])
        elements += [title, Spacer(1, 8), info, Spacer(1, 12)]
        elements.append(_build_tickets_table(participantes))
        doc.build(elements)
        return JSONResponse(content={
            'success': True,
            'path': '/static/tickets_participantes.pdf',
            'total_tickets': len(participantes)
        }, status_code=200)
    except Exception as e:
        logger.error(f"Error generando PDF: {e}")
        return JSONResponse(content={'success': False, 'message': f'No se pudo generar el PDF: {str(e)}'}, status_code=500)

@app.get('/descargar-tickets-pdf')
@login_required
def descargar_tickets_pdf(request: Request):
    path = Path('static') / 'tickets_participantes.pdf'
    if not path.exists():
        # Si no existe, lo generamos al vuelo
        generar_tickets_pdf(request)
    if path.exists():
        return FileResponse(path, filename='tickets_participantes.pdf', media_type='application/pdf')
    return JSONResponse(content={'success': False, 'message': 'No hay PDF generado aún'}, status_code=404)

@app.post('/admin/asignar-tickets')
@login_required
@handle_database_error
async def asignar_tickets(request: Request):
    """Asigna cantidad de tickets a un participante existente."""
    data = await request.json()
    participante_id = data.get('participante_id')
    cantidad_tickets = data.get('cantidad_tickets', 0)
    observaciones = data.get('observaciones', '')
    aprobar = data.get('aprobar', False)
    
    if not participante_id or cantidad_tickets < 1:
        return JSONResponse(content={'success': False, 'message': 'Datos inválidos'}, status_code=400)
    with DatabaseManager() as conn:
        cursor = conn.cursor()
        
        # Obtener datos del participante original
        participante = conn.execute('SELECT * FROM participantes WHERE id = ?', (participante_id,)).fetchone()
        if not participante:
            return JSONResponse(content={'success': False, 'message': 'Participante no encontrado'}, status_code=404)
        # Crear tickets adicionales (cantidad_tickets - 1 porque ya tiene uno)
        tickets_creados = []
        numeros_asignados = []

        for _ in range(cantidad_tickets - 1):
            # Retry logic for race condition protection
            max_retries = 10
            ticket_created = False

            for retry_attempt in range(max_retries):
                try:
                    numero_participacion = _generar_numero_participacion_unico(conn)

                    cursor.execute('''
                        INSERT INTO participantes
                        (numero_participacion, sorteo_id, tipo_documento, numero_documento, nombres, apellidos,
                         whatsapp, departamento, comprobante_path, comprobante_estado,
                         comprobante_observaciones, validado_por)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        RETURNING id
                    ''', (
                        numero_participacion,
                        participante['sorteo_id'],
                        participante['tipo_documento'],
                        participante['numero_documento'],
                        participante['nombres'],
                        participante['apellidos'],
                        participante['whatsapp'],
                        participante['departamento'],
                        participante['comprobante_path'],
                        'aprobado' if aprobar else participante['comprobante_estado'],
                        observaciones if observaciones else participante['comprobante_observaciones'],
                        request.session.get('admin_username', 'Administrador')
                    ))

                    tickets_creados.append(cursor.lastrowid)
                    numeros_asignados.append(numero_participacion)
                    ticket_created = True
                    break  # Success, exit retry loop

                except PsycopgIntegrityError as e:
                    if 'numero_participacion' in str(e) and retry_attempt < max_retries - 1:
                        # Race condition detected, retry with a new number
                        logger.warning(f"Colisión de numero_participacion al asignar tickets (intento {retry_attempt + 1}/{max_retries}), reintentando...")
                        import time
                        time.sleep(0.01 * (retry_attempt + 1))  # Small exponential backoff
                        continue
                    else:
                        # Other integrity error or max retries reached
                        logger.error(f"Error al crear ticket adicional después de {retry_attempt + 1} intentos: {e}")
                        break

            if not ticket_created:
                logger.error(f"No se pudo crear ticket adicional después de {max_retries} intentos")
        
        # Si se debe aprobar, actualizar el ticket original también
        if aprobar and participante['comprobante_estado'] != 'aprobado':
            peru_now = get_peru_time().isoformat()
            cursor.execute('''
                UPDATE participantes 
                SET comprobante_estado = 'aprobado',
                    comprobante_observaciones = ?,
                    validado_por = ?,
                    fecha_validacion = ?,
                    updated_at = ?
                WHERE id = ?
            ''', (observaciones, request.session.get('admin_username', 'Administrador'), peru_now, peru_now, participante_id))
            
            # Registrar en historial
            cursor.execute('''
                INSERT INTO historial_validaciones 
                (participante_id, estado, observaciones, validado_por)
                VALUES (?, ?, ?, ?)
            ''', (participante_id, 'aprobado', observaciones, request.session.get('admin_username', 'Administrador')))
    
    return JSONResponse(content={
        'success': True,
        'message': f'Se asignaron {cantidad_tickets} tickets exitosamente',
        'tickets_creados': len(tickets_creados) + 1,
        'numeros_asignados': numeros_asignados
    }, status_code=200)
        # ====== ✅ NUEVO: GESTIÓN DE SORTEOS MÚLTIPLES ======

@app.get('/admin/sorteos')
@login_required
def listar_sorteos(request: Request):
    """Lista todos los sorteos del sistema."""
    with DatabaseManager() as conn:
        sorteos = conn.execute('''
            SELECT 
                s.*,
                COUNT(DISTINCT CASE WHEN p.comprobante_estado = 'aprobado' THEN p.id END) as total_participantes,
                COUNT(DISTINCT CASE WHEN p.comprobante_estado = 'aprobado' THEN p.id END) as participantes_aprobados,
                COUNT(DISTINCT g.id) as total_ganadores
            FROM sorteos s
            LEFT JOIN participantes p ON s.id = p.sorteo_id AND p.activo IS TRUE
            LEFT JOIN ganadores g ON s.id = g.sorteo_id
            GROUP BY s.id
            ORDER BY s.created_at DESC
        ''').fetchall()
    
    return JSONResponse(content={
        'success': True,
        'sorteos': [dict(s) for s in sorteos]
    }, status_code=200)

@app.get('/admin/sorteo-activo')
@login_required
def obtener_sorteo_activo_api(request: Request):
    """Obtiene información del sorteo activo."""
    sorteo = get_sorteo_activo()
    
    if not sorteo:
        return JSONResponse(content={
            'success': False,
            'message': 'No hay sorteo activo'
        }, status_code=404)
    with DatabaseManager() as conn:
        stats = {
            'total_participantes': conn.execute(
                "SELECT COUNT(*) as count FROM participantes WHERE sorteo_id = ? AND activo IS TRUE AND comprobante_estado = 'aprobado'",
                (sorteo['id'],)
            ).fetchone()['count'],
            'participantes_aprobados': conn.execute(
                "SELECT COUNT(*) as count FROM participantes WHERE sorteo_id = ? AND comprobante_estado = 'aprobado' AND activo IS TRUE",
                (sorteo['id'],)
            ).fetchone()['count'],
            'participantes_pendientes': conn.execute(
                "SELECT COUNT(*) as count FROM participantes WHERE sorteo_id = ? AND comprobante_estado = 'pendiente' AND activo IS TRUE",
                (sorteo['id'],)
            ).fetchone()['count']
        }
    
    return JSONResponse(content={
        'success': True,
        'sorteo': sorteo,
        'stats': stats
    }, status_code=200)

@app.post('/admin/sorteo/{sorteo_id}/cerrar')
@login_required
@handle_database_error
def cerrar_sorteo(request: Request, sorteo_id: int):
    """Cierra un sorteo activo."""
    with DatabaseManager() as conn:
        cursor = conn.cursor()
        
        # Verificar que el sorteo existe y está activo
        sorteo = conn.execute('SELECT * FROM sorteos WHERE id = ?', (sorteo_id,)).fetchone()
        if not sorteo:
            return JSONResponse(content={'success': False, 'message': 'Sorteo no encontrado'}, status_code=404)
        if sorteo['estado'] != 'activo':
            return JSONResponse(content={'success': False, 'message': f'El sorteo ya está en estado: {sorteo["estado"]}'}, status_code=400)
        # Cerrar el sorteo
        peru_now = get_peru_time().isoformat()
        cursor.execute('''
            UPDATE sorteos 
            SET estado = 'cerrado',
                fecha_cierre = ?,
                updated_at = ?
            WHERE id = ?
        ''', (peru_now, peru_now, sorteo_id,))
    
    return JSONResponse(content={
        'success': True,
        'message': f'Sorteo "{sorteo["nombre"]}" cerrado exitosamente'
    }, status_code=200)

@app.post('/admin/sorteo/nuevo')
@login_required
@handle_database_error
async def crear_nuevo_sorteo(request: Request):
    """Crea un nuevo sorteo y lo activa."""
    try:
        data = await request.json()
    except Exception:
        data = {}
    nombre = data.get('nombre', f'Sorteo {datetime.now().strftime("%Y-%m-%d")}')
    descripcion = data.get('descripcion', '')
    
    with DatabaseManager() as conn:
        cursor = conn.cursor()
        
        # Verificar si hay un sorteo activo
        sorteo_activo_actual = conn.execute("SELECT * FROM sorteos WHERE estado = 'activo'").fetchone()
        if sorteo_activo_actual:
            return JSONResponse(content={
                'success': False,
                'message': f'Ya existe un sorteo activo: "{sorteo_activo_actual["nombre"]}". Debe cerrarlo primero.'
            }, status_code=400)
        # Crear nuevo sorteo
        peru_now = get_peru_time().isoformat()
        cursor.execute('''
            INSERT INTO sorteos (nombre, fecha_sorteo, descripcion, estado, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            RETURNING id
        ''', (
            nombre,
            peru_now,
            descripcion,
            'activo',
            peru_now,
            peru_now
        ))
        
        nuevo_sorteo_id = cursor.lastrowid
    
    return JSONResponse(content={
        'success': True,
        'message': f'Nuevo sorteo "{nombre}" creado y activado',
        'sorteo_id': nuevo_sorteo_id
    }, status_code=201)

@app.delete('/admin/sorteo/{sorteo_id}/eliminar')
@login_required
@handle_database_error
def eliminar_sorteo(request: Request, sorteo_id: int):
    """🗑️ NUEVO: Elimina un sorteo de forma segura con validaciones estrictas."""
    with DatabaseManager() as conn:
        cursor = conn.cursor()
        
        # Verificar que el sorteo existe
        sorteo = conn.execute('SELECT * FROM sorteos WHERE id = ?', (sorteo_id,)).fetchone()
        if not sorteo:
            return JSONResponse(content={'success': False, 'message': 'Sorteo no encontrado'}, status_code=404)
        # 🔒 VALIDACIÓN 1: No se puede eliminar el sorteo ACTIVO
        if sorteo['estado'] == 'activo':
            return JSONResponse(content={
                'success': False,
                'message': '❌ No se puede eliminar el sorteo ACTIVO. Primero debes cerrarlo o crear uno nuevo.'
            }, status_code=403)
        # 🔒 VALIDACIÓN 2: Verificar si tiene ganadores
        ganadores = conn.execute(
            'SELECT COUNT(*) as count FROM ganadores WHERE sorteo_id = ?',
            (sorteo_id,)
        ).fetchone()
        
        if ganadores['count'] > 0:
            return JSONResponse(content={
                'success': False,
                'message': f'⚠️ Este sorteo tiene {ganadores["count"]} ganador(es) registrado(s). Por transparencia, no se puede eliminar. Solo puedes marcarlo como cancelado.'
            }, status_code=403)
        # 🔒 VALIDACIÓN 3: Verificar cuántos participantes tiene
        participantes = conn.execute(
            'SELECT COUNT(*) as count FROM participantes WHERE sorteo_id = ? AND activo IS TRUE',
            (sorteo_id,)
        ).fetchone()
        
        # Si tiene participantes, advertir (pero permitir eliminación)
        if participantes['count'] > 0:
            # Los participantes quedarán huérfanos (sorteo_id = NULL)
            cursor.execute(
                'UPDATE participantes SET sorteo_id = NULL WHERE sorteo_id = ?',
                (sorteo_id,)
            )
            logger.info(f'Se desvincularon {participantes["count"]} participantes del sorteo {sorteo_id}')
        
        # ✅ Eliminar el sorteo (CASCADE eliminará ganadores automáticamente por FK)
        cursor.execute('DELETE FROM sorteos WHERE id = ?', (sorteo_id,))
        
        mensaje = f'✅ Sorteo "{sorteo["nombre"]}" eliminado exitosamente.'
        if participantes['count'] > 0:
            mensaje += f' Los {participantes["count"]} participantes han sido desvinculados y podrán asignarse a un nuevo sorteo.'
    
    return JSONResponse(content={
        'success': True,
        'message': mensaje,
        'participantes_desvinculados': participantes['count'] if participantes else 0
    }, status_code=200)

@app.post('/admin/sorteos/limpiar-antiguos')
@login_required
@handle_database_error
def limpiar_sorteos_antiguos(request: Request):
    """🧹 Elimina sorteos antiguos dejando solo los 2 más recientes."""
    with DatabaseManager() as conn:
        cursor = conn.cursor()

        # Obtener todos los sorteos ordenados por fecha de creación (más recientes primero)
        sorteos = conn.execute('''
            SELECT id, nombre, estado, fecha_creacion,
                   (SELECT COUNT(*) FROM ganadores WHERE sorteo_id = sorteos.id) as ganadores_count,
                   (SELECT COUNT(*) FROM participantes WHERE sorteo_id = sorteos.id) as participantes_count
            FROM sorteos
            ORDER BY fecha_creacion DESC
        ''').fetchall()

        total_sorteos = len(sorteos)

        # Si hay 2 o menos sorteos, no hay nada que limpiar
        if total_sorteos <= 2:
            return JSONResponse(content={
                'success': True,
                'message': f'No hay sorteos antiguos que limpiar. Solo tienes {total_sorteos} sorteo(s).',
                'eliminados': 0
            }, status_code=200)
        # Sorteos a mantener (los 2 más recientes)
        sorteos_a_mantener = [s['id'] for s in sorteos[:2]]

        # Sorteos a eliminar (todos excepto los 2 más recientes)
        sorteos_a_eliminar = sorteos[2:]

        eliminados = 0
        no_eliminados = []
        participantes_desvinculados = 0

        for sorteo in sorteos_a_eliminar:
            # Validar: No eliminar si es el sorteo activo
            if sorteo['estado'] == 'activo':
                no_eliminados.append({
                    'nombre': sorteo['nombre'],
                    'razon': 'Es el sorteo ACTIVO'
                })
                continue

            # Validar: No eliminar si tiene ganadores (transparencia)
            if sorteo['ganadores_count'] > 0:
                no_eliminados.append({
                    'nombre': sorteo['nombre'],
                    'razon': f'Tiene {sorteo["ganadores_count"]} ganador(es) registrado(s)'
                })
                continue

            # Desvincular participantes (si los tiene)
            if sorteo['participantes_count'] > 0:
                cursor.execute(
                    'UPDATE participantes SET sorteo_id = NULL WHERE sorteo_id = ?',
                    (sorteo['id'],)
                )
                participantes_desvinculados += sorteo['participantes_count']

            # Eliminar sorteo
            cursor.execute('DELETE FROM sorteos WHERE id = ?', (sorteo['id'],))
            eliminados += 1
            logger.info(f'Sorteo "{sorteo["nombre"]}" (ID: {sorteo["id"]}) eliminado por limpieza automática')

        # Construir mensaje de respuesta
        mensaje_partes = []

        if eliminados > 0:
            mensaje_partes.append(f'✅ Se eliminaron {eliminados} sorteo(s) antiguo(s).')
            if participantes_desvinculados > 0:
                mensaje_partes.append(f'{participantes_desvinculados} participante(s) fueron desvinculados.')

        if no_eliminados:
            mensaje_partes.append(f'⚠️ {len(no_eliminados)} sorteo(s) no se pudieron eliminar:')
            for item in no_eliminados:
                mensaje_partes.append(f'  • {item["nombre"]}: {item["razon"]}')

        if eliminados == 0 and not no_eliminados:
            mensaje_partes.append('No se eliminó ningún sorteo.')

        return JSONResponse(content={
            'success': True,
            'message': ' '.join(mensaje_partes),
            'eliminados': eliminados,
            'no_eliminados': len(no_eliminados),
            'participantes_desvinculados': participantes_desvinculados,
            'sorteos_mantenidos': len(sorteos_a_mantener)
        }, status_code=200)

@app.get('/admin/sorteo/{sorteo_id}/registrar-ganadores')
@login_required
@handle_database_error
def obtener_datos_asignacion_ganadores(request: Request, sorteo_id: int):
    """✅ NUEVO: Obtiene datos necesarios para asignar ganadores manualmente."""
    with DatabaseManager() as conn:
        # Verificar que el sorteo existe
        sorteo = conn.execute('SELECT * FROM sorteos WHERE id = ?', (sorteo_id,)).fetchone()
        if not sorteo:
            return JSONResponse(content={'success': False, 'message': 'Sorteo no encontrado'}, status_code=404)
        # Obtener participantes aprobados del sorteo
        participantes = conn.execute('''
            SELECT * FROM participantes
            WHERE sorteo_id = ?
            AND activo IS TRUE
            AND comprobante_estado = 'aprobado'
            ORDER BY numero_participacion ASC
        ''', (sorteo_id,)).fetchall()
        
        # Obtener premios activos con información de ganadores
        premios = conn.execute('''
            SELECT 
                p.*,
                COALESCE(p.cantidad_ganadores, 1) as cantidad_ganadores,
                COUNT(g.id) as ganadores_asignados
            FROM premios p
            LEFT JOIN ganadores g ON p.id = g.premio_id AND g.sorteo_id = ?
            WHERE p.activo IS TRUE
            GROUP BY p.id
            ORDER BY p.orden ASC
        ''', (sorteo_id,)).fetchall()
        
        # Obtener ganadores existentes
        ganadores_existentes = conn.execute('''
            SELECT 
                g.*,
                p.numero_participacion,
                p.nombres,
                p.apellidos,
                pr.nombre as premio_nombre,
                pr.orden as premio_orden
            FROM ganadores g
            JOIN participantes p ON g.participante_id = p.id
            JOIN premios pr ON g.premio_id = pr.id
            WHERE g.sorteo_id = ?
            ORDER BY pr.orden ASC, g.fecha_ganador ASC
        ''', (sorteo_id,)).fetchall()
    
    return JSONResponse(content={
        'success': True,
        'sorteo': dict(sorteo),
        'participantes': [dict(p) for p in participantes],
        'premios': [dict(p) for p in premios],
        'ganadores_existentes': [dict(g) for g in ganadores_existentes]
    }, status_code=200)

@app.post('/admin/sorteo/{sorteo_id}/registrar-ganadores')
@login_required
@handle_database_error
async def asignar_ganador_manual(request: Request, sorteo_id: int):
    """✅ MEJORADO: Asigna un ganador manualmente - Permite múltiples ganadores por premio."""
    try:
        data = await request.json()
    except Exception:
        data = {}
    numero_ticket = data.get('numero_ticket', '').strip()
    premio_id = data.get('premio_id')

    if not numero_ticket or not premio_id:
        return JSONResponse(content={
            'success': False,
            'message': 'Debe proporcionar el número de ticket y el premio'
        }, status_code=400)
    with DatabaseManager() as conn:
        cursor = conn.cursor()

        # Verificar que el sorteo existe
        sorteo = conn.execute('SELECT * FROM sorteos WHERE id = ?', (sorteo_id,)).fetchone()
        if not sorteo:
            return JSONResponse(content={'success': False, 'message': 'Sorteo no encontrado'}, status_code=404)
        # Buscar el participante por número de ticket
        participante = conn.execute('''
            SELECT * FROM participantes 
            WHERE numero_participacion = ? 
            AND sorteo_id = ?
            AND activo IS TRUE
            AND comprobante_estado = 'aprobado'
        ''', (numero_ticket, sorteo_id)).fetchone()
        
        if not participante:
            return JSONResponse(content={
                'success': False,
                'message': f'No se encontró ningún participante con ticket #{numero_ticket} aprobado en este sorteo'
            }, status_code=404)
        # Verificar que el participante no haya ganado ya en este sorteo
        ya_gano = conn.execute('''
            SELECT * FROM ganadores WHERE sorteo_id = ? AND participante_id = ?
        ''', (sorteo_id, participante['id'])).fetchone()
        
        if ya_gano:
            return JSONResponse(content={
                'success': False,
                'message': f'El participante con ticket #{numero_ticket} ya ganó un premio en este sorteo'
            }, status_code=400)
        # Verificar que el premio existe
        premio = conn.execute('SELECT * FROM premios WHERE id = ? AND activo IS TRUE', (premio_id,)).fetchone()
        if not premio:
            return JSONResponse(content={'success': False, 'message': 'Premio no encontrado'}, status_code=404)
        # ✅ CORREGIDO: Contar cuántos ganadores ya tiene este premio en este sorteo
        ganadores_actuales = conn.execute('''
            SELECT COUNT(*) as count FROM ganadores
            WHERE sorteo_id = ? AND premio_id = ?
        ''', (sorteo_id, premio_id)).fetchone()['count']

        # ✅ CORREGIDO: Verificar si el premio tiene límite de ganadores
        # sqlite3.Row se accede con corchetes, no con .get()
        # Manejo robusto de NULL o valores inválidos
        try:
            cantidad_maxima = int(premio['cantidad_ganadores']) if premio['cantidad_ganadores'] is not None else 1
            if cantidad_maxima < 1:
                cantidad_maxima = 1
        except (ValueError, TypeError):
            # Si hay error de conversión, usar 1 por defecto
            cantidad_maxima = 1
            logger.warning(f'Premio {premio_id} tiene cantidad_ganadores inválido: {premio["cantidad_ganadores"]}')
        
        # ✅ VALIDAR: Solo si el premio tiene MÚLTIPLES GANADORES, verificar límite
        if cantidad_maxima > 1:
            if ganadores_actuales >= cantidad_maxima:
                return JSONResponse(content={
                    'success': False,
                    'message': f'El premio "{premio["nombre"]}" ya alcanzó su límite de {cantidad_maxima} ganador(es). Actualmente tiene {ganadores_actuales} ganador(es).'
                }, status_code=400)
        else:
            # Si es un solo ganador, no se puede asignar otro
            if ganadores_actuales > 0:
                return JSONResponse(content={
                    'success': False,
                    'message': f'El premio "{premio["nombre"]}" ya tiene un ganador asignado'
                }, status_code=400)
        # Registrar el ganador
        cursor.execute('''
            INSERT INTO ganadores (sorteo_id, participante_id, premio_id)
            VALUES (?, ?, ?)
            RETURNING id
        ''', (sorteo_id, participante['id'], premio_id))

        # 🆕 VERIFICAR después del INSERT por si hubo race condition
        nuevos_ganadores_count = conn.execute('''
            SELECT COUNT(*) as count FROM ganadores
            WHERE sorteo_id = ? AND premio_id = ?
        ''', (sorteo_id, premio_id)).fetchone()['count']

        # ⚠️ Si se excedió el límite por race condition, advertir en log
        if cantidad_maxima > 1 and nuevos_ganadores_count > cantidad_maxima:
            logger.warning(
                f'⚠️ RACE CONDITION: Premio {premio["nombre"]} (ID:{premio_id}) excedió límite. '
                f'Esperado: {cantidad_maxima}, Actual: {nuevos_ganadores_count}'
            )
        
        ganador_info = {
            'ganador_id': cursor.lastrowid,
            'numero_ticket': participante['numero_participacion'],
            'nombre_completo': f"{participante['nombres']} {participante['apellidos']}",
            'documento': f"{participante['tipo_documento']}: {participante['numero_documento']}",
            'departamento': participante['departamento'],
            'whatsapp': participante['whatsapp'],
            'premio': premio['nombre'],
            'premio_id': premio_id,
            'ganadores_actuales': nuevos_ganadores_count,
            'ganadores_maximos': cantidad_maxima
        }
    
    mensaje_extra = ""
    if nuevos_ganadores_count < cantidad_maxima:
        faltan = cantidad_maxima - nuevos_ganadores_count
        mensaje_extra = f" (Faltan {faltan} ganador(es) más para este premio)"
    elif nuevos_ganadores_count == cantidad_maxima:
        mensaje_extra = " (Premio completado ✅)"
    
    return JSONResponse(content={
        'success': True,
        'message': f'¡Ganador asignado! Ticket #{numero_ticket} ganó {premio["nombre"]}{mensaje_extra}',
        'ganador': ganador_info
    }, status_code=201)

@app.get('/admin/sorteo/{sorteo_id}/ganadores')
@login_required
def ver_ganadores_sorteo(request: Request, sorteo_id: int):
    """Lista todos los ganadores de un sorteo específico."""
    with DatabaseManager() as conn:
        ganadores = conn.execute('''
            SELECT 
                g.*,
                p.numero_participacion,
                p.nombres,
                p.apellidos,
                p.tipo_documento,
                p.numero_documento,
                p.whatsapp,
                p.departamento,
                pr.nombre as premio_nombre,
                pr.descripcion as premio_descripcion,
                pr.orden as premio_orden,
                COALESCE(pr.cantidad_ganadores, 1) as premio_cantidad_ganadores
            FROM ganadores g
            JOIN participantes p ON g.participante_id = p.id
            JOIN premios pr ON g.premio_id = pr.id
            WHERE g.sorteo_id = ?
            ORDER BY pr.orden ASC, g.fecha_ganador ASC
        ''', (sorteo_id,)).fetchall()
    
    return JSONResponse(content={
        'success': True,
        'ganadores': [dict(g) for g in ganadores]
    }, status_code=200)

@app.get('/sorteos')
def lista_sorteos_publico(request: Request):
    """Lista pública de todos los sorteos disponibles."""
    with DatabaseManager() as conn:
        sorteos = conn.execute('''
            SELECT 
                s.*,
                COUNT(DISTINCT CASE WHEN p.comprobante_estado = 'aprobado' THEN p.id END) as total_participantes,
                COUNT(DISTINCT CASE WHEN p.comprobante_estado = 'aprobado' THEN p.id END) as participantes_aprobados,
                COUNT(DISTINCT g.id) as total_ganadores
            FROM sorteos s
            LEFT JOIN participantes p ON s.id = p.sorteo_id AND p.activo IS TRUE
            LEFT JOIN ganadores g ON s.id = g.sorteo_id
            GROUP BY s.id
            ORDER BY s.created_at DESC
        ''').fetchall()
    
    return templates.TemplateResponse('lista_sorteos.html', template_ctx(request, sorteos=[dict(s) for s in sorteos]))

@app.get('/sorteos/{sorteo_id}/ganadores/ver')
def ver_ganadores_sorteo_publico(request: Request, sorteo_id: int):
    """Vista pública de ganadores de un sorteo específico."""
    with DatabaseManager() as conn:
        # Obtener información del sorteo
        sorteo = conn.execute('SELECT * FROM sorteos WHERE id = ?', (sorteo_id,)).fetchone()
        if not sorteo:
            return templates.TemplateResponse('404.html', template_ctx(request), status_code=404)

        # ✅ MEJORADO: Obtener ganadores con nombres completos
        ganadores = conn.execute('''
            SELECT 
                p.numero_participacion as ticket,
                p.nombres,
                p.apellidos,
                pr.nombre as premio_nombre,
                pr.descripcion as premio_descripcion,
                pr.orden as premio_orden,
                g.fecha_ganador
            FROM ganadores g
            JOIN participantes p ON g.participante_id = p.id
            JOIN premios pr ON g.premio_id = pr.id
            WHERE g.sorteo_id = ?
            ORDER BY pr.orden ASC
        ''', (sorteo_id,)).fetchall()
        
    return templates.TemplateResponse(
        'ganadores_sorteo.html',
        template_ctx(request, sorteo=dict(sorteo), ganadores=[dict(g) for g in ganadores]),
    )

@app.post('/admin/registrar-compra-adicional')
@login_required
@handle_database_error
async def registrar_compra_adicional(request: Request):
    """Registra una compra adicional para un cliente existente."""
    form = await request.form()
    numero_documento = (form.get('numero_documento') or '').strip()
    try:
        cantidad_tickets = int(form.get('cantidad_tickets') or 0)
    except ValueError:
        cantidad_tickets = 0
    observaciones = (form.get('observaciones') or '').strip()
    aprobar = (form.get('aprobar') or 'false').lower() == 'true'
    
    if not numero_documento or cantidad_tickets < 1:
        return JSONResponse(content={'success': False, 'message': 'Datos inválidos'}, status_code=400)

    # Manejar archivo de comprobante
    comprobante_path = None
    upload = form.get('comprobante')
    if upload is not None and hasattr(upload, 'filename') and getattr(upload, 'filename', None):
        comprobante_path = await FileManager.save_upload_file(upload)
        if not comprobante_path:
            return JSONResponse(content={'success': False, 'message': 'Error al procesar el comprobante'}, status_code=400)
    with DatabaseManager() as conn:
        cursor = conn.cursor()
        
        # Obtener datos del cliente existente
        cliente = conn.execute('''
            SELECT * FROM participantes 
            WHERE numero_documento = ? 
            ORDER BY fecha_registro DESC 
            LIMIT 1
        ''', (numero_documento,)).fetchone()
        
        if not cliente:
            return JSONResponse(content={'success': False, 'message': 'Cliente no encontrado'}, status_code=404)
        # Crear nuevos tickets
        tickets_creados = []
        numeros_asignados = []

        for _ in range(cantidad_tickets):
            # Retry logic for race condition protection
            max_retries = 10
            ticket_created = False

            for retry_attempt in range(max_retries):
                try:
                    numero_participacion = _generar_numero_participacion_unico(conn)

                    # ✅ NUEVO: Obtener sorteo activo
                    sorteo_activo = get_sorteo_activo()
                    if not sorteo_activo:
                        return JSONResponse(content={'success': False, 'message': 'No hay sorteo activo'}, status_code=400)

                    peru_now = get_peru_time().isoformat() if aprobar else None
                    cursor.execute('''
                        INSERT INTO participantes
                        (numero_participacion, sorteo_id, tipo_documento, numero_documento, nombres, apellidos,
                         whatsapp, departamento, comprobante_path, comprobante_estado,
                         comprobante_observaciones, validado_por, fecha_validacion, fecha_registro, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        RETURNING id
                    ''', (
                        numero_participacion,
                        sorteo_activo['id'],
                        cliente['tipo_documento'],
                        cliente['numero_documento'],
                        cliente['nombres'],
                        cliente['apellidos'],
                        cliente['whatsapp'],
                        cliente['departamento'],
                        comprobante_path,
                        'aprobado' if aprobar else 'pendiente',
                        observaciones,
                        request.session.get('admin_username', 'Administrador') if aprobar else None,
                        peru_now,
                        get_peru_time().isoformat(),
                        get_peru_time().isoformat(),
                        get_peru_time().isoformat()
                    ))

                    ticket_id = cursor.lastrowid
                    tickets_creados.append(ticket_id)
                    numeros_asignados.append(numero_participacion)

                    # Si se aprueba automáticamente, registrar en historial
                    if aprobar:
                        cursor.execute('''
                            INSERT INTO historial_validaciones
                            (participante_id, estado, observaciones, validado_por)
                            VALUES (?, ?, ?, ?)
                        ''', (ticket_id, 'aprobado', f'Compra adicional - {observaciones}',
                              request.session.get('admin_username', 'Administrador')))

                    ticket_created = True
                    break  # Success, exit retry loop

                except PsycopgIntegrityError as e:
                    if 'numero_participacion' in str(e) and retry_attempt < max_retries - 1:
                        # Race condition detected, retry with a new number
                        logger.warning(f"Colisión de numero_participacion al registrar compra adicional (intento {retry_attempt + 1}/{max_retries}), reintentando...")
                        import time
                        time.sleep(0.01 * (retry_attempt + 1))  # Small exponential backoff
                        continue
                    else:
                        # Other integrity error or max retries reached
                        logger.error(f"Error al crear ticket después de {retry_attempt + 1} intentos: {e}")
                        break

            if not ticket_created:
                logger.error(f"No se pudo crear ticket después de {max_retries} intentos")
    
    return JSONResponse(content={
        'success': True,
        'message': f'Se registraron {len(tickets_creados)} tickets exitosamente para {cliente["nombres"]} {cliente["apellidos"]}',
        'tickets_creados': tickets_creados,
        'numeros_asignados': numeros_asignados
    }, status_code=201)

# ====== ✅ NUEVO: EDICIÓN DE PARTICIPANTES CON AUDITORÍA ======

@app.put('/admin/participante/{participante_id}/editar')
@login_required
@handle_database_error
async def editar_participante(request: Request, participante_id: int):
    """Edita los datos de un participante con auditoría completa."""
    data = await request.json()
    
    if not data:
        return JSONResponse(content={'success': False, 'message': 'No se recibieron datos para editar'}, status_code=400)
    with DatabaseManager() as conn:
        cursor = conn.cursor()
        
        # 🔒 Verificar que el participante existe
        participante = conn.execute(
            'SELECT * FROM participantes WHERE id = ?', 
            (participante_id,)
        ).fetchone()
        
        if not participante:
            return JSONResponse(content={'success': False, 'message': 'Participante no encontrado'}, status_code=404)
        # 🔒 Verificar que no es ganador
        es_ganador = conn.execute(
            'SELECT COUNT(*) as count FROM ganadores WHERE participante_id = ?',
            (participante_id,)
        ).fetchone()['count']
        
        if es_ganador > 0:
            return JSONResponse(content={
                'success': False, 
                'message': '❌ No se puede editar un participante que ya es ganador. Esto compromete la transparencia del sorteo.'
            }, status_code=403)
        # 🔒 Verificar que el sorteo no está cerrado o completado
        if participante['sorteo_id']:
            sorteo = conn.execute(
                'SELECT estado FROM sorteos WHERE id = ?',
                (participante['sorteo_id'],)
            ).fetchone()
            
            if sorteo and sorteo['estado'] in ['cerrado', 'completado']:
                return JSONResponse(content={
                    'success': False,
                    'message': f'❌ No se puede editar participantes de sorteos {sorteo["estado"]}s. El sorteo ya finalizó.'
                }, status_code=403)
        # 🔒 Si el comprobante está aprobado, solo permitir editar ciertos campos
        campos_permitidos_aprobados = ['whatsapp', 'departamento']
        if participante['comprobante_estado'] == 'aprobado':
            # Verificar que solo se editen campos permitidos
            campos_a_editar = [k for k in data.keys() if k not in ['observaciones'] and data.get(k) != participante[k]]
            campos_no_permitidos = [c for c in campos_a_editar if c not in campos_permitidos_aprobados]
            
            if campos_no_permitidos:
                return JSONResponse(content={
                    'success': False,
                    'message': f'⚠️ Este participante tiene comprobante APROBADO. Solo puedes editar: WhatsApp y Departamento. No puedes modificar: {", ".join(campos_no_permitidos)}'
                }, status_code=403)
        # 📝 Detectar y registrar cambios
        cambios = []
        campos_editables = {
            'tipo_documento': 'Tipo de Documento',
            'numero_documento': 'Número de Documento',
            'nombres': 'Nombres',
            'apellidos': 'Apellidos',
            'whatsapp': 'WhatsApp',
            'departamento': 'Departamento'
        }
        
        for campo, nombre_campo in campos_editables.items():
            if campo in data and str(data[campo]).strip() != str(participante[campo]):
                # Validaciones específicas por campo
                if campo == 'numero_documento':
                    if not data[campo].isdigit() or len(data[campo]) < 7:
                        return JSONResponse(content={
                            'success': False,
                            'message': f'Número de documento inválido (mínimo 7 dígitos)'
                        }, status_code=400)
                if campo == 'tipo_documento' and data[campo] not in ['DNI', 'CE']:
                    return JSONResponse(content={
                        'success': False,
                        'message': 'Tipo de documento inválido. Debe ser DNI o CE'
                    }, status_code=400)
                if campo == 'whatsapp':
                    if not re.match(r'^[\d\s\-\(\)\+]+$', data[campo]):
                        return JSONResponse(content={
                            'success': False,
                            'message': 'Número de WhatsApp inválido'
                        }, status_code=400)
                cambios.append({
                    'campo': campo,
                    'nombre_campo': nombre_campo,
                    'valor_anterior': participante[campo],
                    'valor_nuevo': str(data[campo]).strip().title() if campo in ['nombres', 'apellidos'] else str(data[campo]).strip()
                })
        
        if not cambios:
            return JSONResponse(content={'success': False, 'message': 'No hay cambios para guardar'}, status_code=400)
        # 💾 Actualizar datos
        update_parts = []
        params = []
        for cambio in cambios:
            update_parts.append(f"{cambio['campo']} = ?")
            params.append(cambio['valor_nuevo'])
        
        params.extend([get_peru_time().isoformat(), participante_id])
        
        cursor.execute(f'''
            UPDATE participantes 
            SET {', '.join(update_parts)}, updated_at = ?
            WHERE id = ?
        ''', params)
        
        # 📊 Registrar cada cambio en el historial de auditoría
        observaciones = data.get('observaciones', 'Corrección de datos del participante')
        editado_por = request.session.get('admin_username', 'Administrador')
        fecha_edicion = get_peru_time().isoformat()
        
        for cambio in cambios:
            cursor.execute('''
                INSERT INTO historial_ediciones 
                (participante_id, campo_editado, valor_anterior, valor_nuevo, editado_por, fecha_edicion, observaciones)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                participante_id,
                cambio['nombre_campo'],
                cambio['valor_anterior'],
                cambio['valor_nuevo'],
                editado_por,
                fecha_edicion,
                observaciones
            ))
        
        # Obtener datos actualizados
        participante_actualizado = conn.execute(
            'SELECT * FROM participantes WHERE id = ?',
            (participante_id,)
        ).fetchone()
    
    return JSONResponse(content={
        'success': True,
        'message': f'✅ Participante actualizado exitosamente. Se modificaron {len(cambios)} campo(s).',
        'cambios': cambios,
        'participante': dict(participante_actualizado)
    }, status_code=200)

@app.get('/admin/participante/{participante_id}/historial-ediciones')
@login_required
def ver_historial_ediciones_participante(request: Request, participante_id: int):
    """Ver todo el historial de ediciones de un participante específico."""
    with DatabaseManager() as conn:
        # Verificar que el participante existe
        participante = conn.execute(
            'SELECT * FROM participantes WHERE id = ?',
            (participante_id,)
        ).fetchone()
        
        if not participante:
            return JSONResponse(content={'success': False, 'message': 'Participante no encontrado'}, status_code=404)
        # Obtener historial de ediciones
        ediciones = conn.execute('''
            SELECT * FROM historial_ediciones
            WHERE participante_id = ?
            ORDER BY fecha_edicion DESC
        ''', (participante_id,)).fetchall()
    
    return JSONResponse(content={
        'success': True,
        'participante': dict(participante),
        'historial_ediciones': [dict(e) for e in ediciones],
        'total_ediciones': len(ediciones)
    }, status_code=200)

@app.get('/admin/historial-ediciones-general')
@login_required
def ver_historial_ediciones_general(request: Request):
    """Ver todas las ediciones realizadas en el sistema (auditoría general)."""
    page = qp_int(request, 'page', 1)
    per_page = qp_int(request, 'per_page', 50)
    filtro_editor = qp_str(request, 'editor', '')
    fecha_desde = qp_str(request, 'fecha_desde', '')
    fecha_hasta = qp_str(request, 'fecha_hasta', '')
    
    with DatabaseManager() as conn:
        query = '''
            SELECT 
                h.*,
                p.numero_participacion,
                p.nombres || ' ' || p.apellidos as participante_nombre,
                p.tipo_documento,
                p.numero_documento
            FROM historial_ediciones h
            JOIN participantes p ON h.participante_id = p.id
            WHERE 1=1
        '''
        params = []
        
        if filtro_editor:
            query += ' AND h.editado_por = ?'
            params.append(filtro_editor)
        if fecha_desde:
            query += ' AND (h.fecha_edicion::date) >= ?::date'
            params.append(fecha_desde)
        if fecha_hasta:
            query += ' AND (h.fecha_edicion::date) <= ?::date'
            params.append(fecha_hasta)
        
        # Contar total
        count_query = f'SELECT COUNT(*) as total FROM ({query})'
        total = conn.execute(count_query, params).fetchone()['total']
        
        # Obtener ediciones paginadas
        query += ' ORDER BY h.fecha_edicion DESC LIMIT ? OFFSET ?'
        params.extend([per_page, (page - 1) * per_page])
        ediciones = conn.execute(query, params).fetchall()
        
        # Obtener lista de editores para filtros
        editores = conn.execute(
            'SELECT DISTINCT editado_por FROM historial_ediciones ORDER BY editado_por'
        ).fetchall()
    
    return JSONResponse(content={
        'success': True,
        'ediciones': [dict(e) for e in ediciones],
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page,
        'editores': [e['editado_por'] for e in editores]
    }, status_code=200)

@app.get('/admin/exportar-historial-ediciones')
@login_required
def exportar_historial_ediciones(request: Request):
    """Exporta el historial completo de ediciones a Excel."""
    try:
        with DatabaseManager() as conn:
            ediciones = conn.execute('''
                SELECT 
                    h.id,
                    p.numero_participacion as ticket,
                    p.nombres || ' ' || p.apellidos as participante,
                    p.tipo_documento || ': ' || p.numero_documento as documento,
                    h.campo_editado,
                    h.valor_anterior,
                    h.valor_nuevo,
                    h.editado_por,
                    h.fecha_edicion,
                    h.observaciones
                FROM historial_ediciones h
                JOIN participantes p ON h.participante_id = p.id
                ORDER BY h.fecha_edicion DESC
            ''').fetchall()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial de Ediciones"
        
        # Headers
        headers = ['ID', 'Ticket', 'Participante', 'Documento', 'Campo Editado', 
                  'Valor Anterior', 'Valor Nuevo', 'Editado Por', 'Fecha', 'Observaciones']
        ws.append(headers)
        
        # Estilo de headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Datos
        for e in ediciones:
            row = [
                e['id'], e['ticket'], e['participante'], e['documento'],
                e['campo_editado'], e['valor_anterior'], e['valor_nuevo'],
                e['editado_por'],
                (e['fecha_edicion'].strftime('%Y-%m-%d %H:%M:%S')
                 if hasattr(e['fecha_edicion'], 'strftime')
                 else (str(e['fecha_edicion'])[:19] if e['fecha_edicion'] else '')),
                e['observaciones'] or 'Sin observaciones'
            ]
            ws.append(row)
        
        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        filename = f"historial_ediciones_{get_peru_time().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = Path('static') / filename
        wb.save(filepath)
        
        return FileResponse(filepath, filename=filename, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f"Error al exportar historial de ediciones: {e}")
        return JSONResponse(content={'success': False, 'message': 'Error al exportar historial de ediciones'}, status_code=500)

# ✅ 3. RUTA NUEVA: Vista pública de ganadores por sorteo (SIN datos sensibles)
@app.get('/sorteos/{sorteo_id}/ganadores')
def ver_ganadores_publico(request: Request, sorteo_id: int):
    """Vista pública de ganadores del sorteo - SIN datos sensibles"""
    
    with DatabaseManager() as conn:
        # Verificar que el sorteo existe
        sorteo = conn.execute('SELECT * FROM sorteos WHERE id = ?', (sorteo_id,)).fetchone()
        if not sorteo:
            return JSONResponse(content={'success': False, 'message': 'Sorteo no encontrado'}, status_code=404)
        # Obtener ganadores (SOLO número de ticket/serial y premio - SIN DNI/CORREO)
        ganadores = conn.execute('''
            SELECT 
                p.numero_participacion as ticket,
                pr.nombre as premio_nombre,
                pr.descripcion as premio_descripcion,
                pr.orden as premio_orden,
                g.fecha_ganador,
                s.nombre as sorteo_nombre,
                s.fecha_sorteo,
                s.estado as sorteo_estado
            FROM ganadores g
            JOIN participantes p ON g.participante_id = p.id
            JOIN premios pr ON g.premio_id = pr.id
            JOIN sorteos s ON g.sorteo_id = s.id
            WHERE g.sorteo_id = ?
            ORDER BY pr.orden ASC
        ''', (sorteo_id,)).fetchall()
        
        sorteo_info = {
            'id': sorteo['id'],
            'nombre': sorteo['nombre'],
            'fecha_sorteo': sorteo['fecha_sorteo'],
            'estado': sorteo['estado'],
            'descripcion': sorteo['descripcion']
        }
        
        return JSONResponse(content={
            'success': True,
            'sorteo': sorteo_info,
            'ganadores': [dict(g) for g in ganadores],
            'total_ganadores': len(ganadores)
        }, status_code=200)

def _ofuscar_documento(numero: str) -> str:
    if not numero:
        return ''
    s = str(numero)
    if len(s) <= 4:
        return '*' * len(s)
    return f"{s[:2]}{'*' * (len(s) - 4)}{s[-2:]}"


def _ofuscar_nombre(valor: str) -> str:
    if not valor:
        return ''
    partes = []
    for palabra in str(valor).split():
        if len(palabra) <= 1:
            partes.append(palabra + '*')
        else:
            partes.append(palabra[0] + '*' * (len(palabra) - 1))
    return ' '.join(partes)


@app.get('/participantes-vivo')
def participantes_vivo_view(request: Request):
    """Vista pública: participantes aprobados del sorteo activo, con datos ofuscados."""
    sorteo_activo = get_sorteo_activo()
    return templates.TemplateResponse(
        'participantes_vivo.html',
        template_ctx(request, sorteo_activo=sorteo_activo),
    )


@app.get('/api/participantes-vivo')
def api_participantes_vivo(request: Request):
    """Feed JSON para la vista en vivo. Público, sólo tickets aprobados del sorteo activo."""
    page = max(1, qp_int(request, 'page', 1))
    per_page = qp_int(request, 'per_page', 20)
    if per_page < 5 or per_page > 100:
        per_page = 20
    search = qp_str(request, 'search', '').strip()

    sorteo_activo = get_sorteo_activo()
    if not sorteo_activo:
        return JSONResponse(content={
            'success': True,
            'sorteo': None,
            'participantes': [],
            'total': 0,
            'page': page,
            'per_page': per_page,
            'pages': 0,
        }, status_code=200)

    where = [
        "activo IS TRUE",
        "comprobante_estado = 'aprobado'",
        "sorteo_id = ?",
    ]
    params: list = [sorteo_activo['id']]
    if search:
        where.append("(numero_participacion LIKE ? OR numero_documento LIKE ?)")
        like = f"%{search}%"
        params.extend([like, like])
    where_sql = " AND ".join(where)

    with DatabaseManager() as conn:
        total = conn.execute(
            f"SELECT COUNT(*) as c FROM participantes WHERE {where_sql}",
            params,
        ).fetchone()['c']
        offset = (page - 1) * per_page
        rows = conn.execute(
            f'''
            SELECT numero_participacion, numero_documento, tipo_documento,
                   nombres, apellidos, departamento, fecha_registro
            FROM participantes
            WHERE {where_sql}
            ORDER BY fecha_registro DESC, id DESC
            LIMIT ? OFFSET ?
            ''',
            params + [per_page, offset],
        ).fetchall()

    participantes = [{
        'numero_participacion': r['numero_participacion'],
        'tipo_documento': r['tipo_documento'],
        'documento': _ofuscar_documento(r['numero_documento']),
        'nombres': _ofuscar_nombre(r['nombres']),
        'apellidos': _ofuscar_nombre(r['apellidos']),
        'departamento': r['departamento'],
        'fecha_registro': r['fecha_registro'],
    } for r in rows]

    return JSONResponse(content={
        'success': True,
        'sorteo': {
            'id': sorteo_activo['id'],
            'nombre': sorteo_activo['nombre'],
            'fecha_sorteo': sorteo_activo['fecha_sorteo'],
            'estado': sorteo_activo['estado'],
        },
        'participantes': participantes,
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page if total else 0,
    }, status_code=200)


app.mount("/static", StaticFiles(directory="static"), name="static")


# Función principal
def main():
    init_database()
    if not config.DEBUG:
        logging.basicConfig(level=logging.INFO)
    print(f"""
    SISTEMA DE SORTEOS RAPI RIFA
    =====================================
    Servidor: http://{config.HOST}:{config.PORT}
    Panel admin: http://{config.HOST}:{config.PORT}/admin
    =====================================
    """)
    import uvicorn

    uvicorn.run("app:app", host=config.HOST, port=config.PORT, reload=config.DEBUG)

if __name__ == '__main__':
    main()